import os, io, hmac, struct, time, base64, hashlib, json, logging, threading
import httpx
from datetime import datetime
from http.server import HTTPServer, BaseHTTPRequestHandler
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from telegram import Update, ReplyKeyboardMarkup, ReplyKeyboardRemove, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Application, CommandHandler, MessageHandler, ConversationHandler, ContextTypes, filters, CallbackQueryHandler

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

BOT_TOKEN       = os.getenv("BOT_TOKEN")
GEMINI_API_KEY  = os.getenv("GEMINI_API_KEY")
MAPILLARY_TOKEN = os.getenv("MAPILLARY_TOKEN")
TOTP_SECRET     = os.getenv("TOTP_SECRET")
DOMINIO         = os.getenv("DOMINIO_EMAIL", "telconet.ec")
GEMINI_URL      = "https://generativelanguage.googleapis.com/v1beta/models/gemini-1.5-flash:generateContent?key=" + (GEMINI_API_KEY or "")

USUARIOS_AUTENTICADOS = set()
RUTAS_GUARDADAS = {}

# MODO PRUEBA: solo se generan REPORTES_DE_RECORRIDOS y FOTOS_ANEXAS_AL_REPORTE.
# Cuando estas 2 pestanas esten validadas, cambiar a True para reactivar
# MANGAS, INVENTARIO DE HILOS EN NODO, Checklist CIU y Checklists MPRIU.
GENERAR_HOJAS_EXTRA = False

(ESPERANDO_TOTP, MENU_PRINCIPAL, NOMBRE_RUTA, CODIGO_CUADRILLA, NODO_INICIAL, NODO_FINAL,
 LIDER, AYUDANTE, COORDINADOR, PLACA, DISTANCIA,
 CIU_HERRAMIENTAS, CIU_EQUIPOS, CIU_MATERIALES,
 NOVEDADES_AUTO, TAREA_PENDIENTE, FOTO_ANTES, FOTO_DESPUES, OBSERVACIONES,
 MPRIU_CHECK, PREGUNTA_MANGAS, PREGUNTA_HILOS,
 MANGA_NOMBRE, MANGA_COORDS, MANGA_OBS, HILO_ODF, HILO_DATOS,
 NUEVA_RUTA_NOMBRE, NUEVA_RUTA_VIDEO,
 GENERAR_CONFIRMAR, GENERAR_NOMBRE_RUTA, GENERAR_CUADRILLA,
 GENERAR_NODO_INI, GENERAR_NODO_FIN, GENERAR_LIDER,
 GENERAR_AYUDANTE, GENERAR_COORDINADOR, GENERAR_PLACA,
 GENERAR_DISTANCIA, GENERAR_FECHA, GENERAR_HORA_INI,
 GENERAR_HORA_FIN, GENERAR_NOVEDADES,
 TAB_MENU, TAB_CIU_HERR, TAB_CIU_EQUI, TAB_CIU_MATE,
 TAB_MPRIU, TAB_REPORTES, TAB_NOVEDADES_IA,
 VIDEO_BASE_NOMBRE, VIDEO_BASE_UPLOAD) = range(52)

NOVEDADES_MPRIU = [
    "HERRAJES EN MAL ESTADO.", "FALTA DE HERRAJES.", "POSTES EN MAL ESTADO.",
    "POSTE(S) CAMBIADO(S).", "POSTES POR INSTALAR.", "POSTE NUEVO INSTALADO \u2013 TN.",
    "POSTE NUEVO INSTALADO \u2013 EMPRESAS EL\u00c9CTRICAS.", "POSTES INCLINADOS.",
    "RETENIDA(S) EN MAL ESTADO.", "RETENIDA(S) CORTADA(S).", "VANOS POR RETEMPLAR.",
    "MANGAS SUELTAS.", "MANGAS ABIERTAS/DA\u00d1ADAS.", "RESERVAS SUELTAS.",
    "CRUCES DE V\u00cdAS BAJOS.", "VEGETACI\u00d3N SOBRE FIBRA/MANGA.", "LOCALIZACI\u00d3N DE MANGA.",
    "DOCUMENTACI\u00d3N UNIFILAR DE HILOS.", "L\u00cdNEA EL\u00c9CTRICA EN MAL ESTADO.",
    "REGENERACI\u00d3N URBANA.", "AMPLIACI\u00d3N DE V\u00cdA.", "CABLE LASTIMADO.",
    "FIBRA INSTALADA INCORRECTAMENTE SOBRE MORDAZA.", "POZO SIN TAPA O EN MAL ESTADO.",
    "REPINTADO DE POZO.", "REPINTADO DE POSTE.", "ELEMENTOS SIN ETIQUETAS ACR\u00cdLICAS.",
    "RIESGO DE DERRUMBE O DESLAVE.", "RIESGO DE INUNDACIONES.", "RIESGO DE INCENDIO.",
    "NO SE REGISTRAN NOVEDADES DURANTE LA INSPECCI\u00d3N.",
]

SIN_NOV_MOTIVO  = "NO SE REGISTRAN NOVEDADES DURANTE LA INSPECCI\u00d3N."
SIN_NOV_REMEDIO = "NO SE ENCUENTRAN NOVEDADES QUE SIGNIFIQUEN RIESGOS EN EL CABLE DE LA RED INTERURBANO."

REMEDIOS = {
    "HERRAJES EN MAL ESTADO.": "REALIZAR EL REEMPLAZO INMEDIATO DEL HERRAJE AFECTADO, GARANTIZANDO LA CORRECTA SUJECI\u00d3N DEL CABLE Y LA ESTABILIDAD MEC\u00c1NICA DEL TENDIDO.",
    "FALTA DE HERRAJES.": "INSTALAR LOS HERRAJES CONFORME A LA NORMATIVA T\u00c9CNICA, ASEGURANDO LA CORRECTA FIJACI\u00d3N DEL CABLE AL POSTE.",
    "POSTES EN MAL ESTADO.": "DOCUMENTAR MEDIANTE REGISTRO FOTOGR\u00c1FICO Y COORDENADAS, Y REPORTAR PARA GESTIONAR EL REEMPLAZO DEL POSTE CON LA ENTIDAD RESPONSABLE.",
    "POSTE(S) CAMBIADO(S).": "INSTALAR LOS HERRAJES NECESARIOS Y ASEGURAR CORRECTAMENTE EL CABLE AL NUEVO POSTE. DOCUMENTAR EL CAMBIO PARA ACTUALIZACI\u00d3N DE INVENTARIO.",
    "POSTES POR INSTALAR.": "DOCUMENTAR LA UBICACI\u00d3N EXACTA Y REPORTAR PARA LA COORDINACI\u00d3N E INSTALACI\u00d3N DEL NUEVO POSTE REQUERIDO.",
    "POSTE NUEVO INSTALADO \u2013 TN.": "DOCUMENTAR, ETIQUETAR CON C\u00d3DIGO DE IDENTIFICACI\u00d3N Y APLICAR PINTURA DE SE\u00d1ALIZACI\u00d3N CONFORME A EST\u00c1NDARES OPERATIVOS.",
    "POSTE NUEVO INSTALADO \u2013 EMPRESAS EL\u00c9CTRICAS.": "DOCUMENTAR, COLOCAR ETIQUETA ACR\u00cdLICA Y ASEGURAR EL CABLE DE FIBRA \u00d3PTICA CONFORME A LA NORMATIVA T\u00c9CNICA VIGENTE, RESPETANDO LAS DISTANCIAS DE SEGURIDAD EL\u00c9CTRICA.",
    "POSTES INCLINADOS.": "DOCUMENTAR MEDIANTE REGISTRO FOTOGR\u00c1FICO Y COORDENADAS, Y REPORTAR PARA GESTIONAR EL APLOME DEL POSTE CON EL CONTRATISTA.",
    "RETENIDA(S) EN MAL ESTADO.": "DOCUMENTAR MEDIANTE REGISTRO FOTOGR\u00c1FICO Y COORDENADAS, Y REPORTAR PARA GESTIONAR LA CORRECCI\u00d3N CON EL CONTRATISTA.",
    "RETENIDA(S) CORTADA(S).": "DOCUMENTAR MEDIANTE REGISTRO FOTOGR\u00c1FICO Y COORDENADAS, Y REPORTAR PARA GESTIONAR LA CORRECCI\u00d3N CON EL CONTRATISTA.",
    "VANOS POR RETEMPLAR.": "REALIZAR EL RETEMPLADO DEL CABLE PARA RESTABLECER LA TENSI\u00d3N ADECUADA Y EVITAR RIESGOS DE DA\u00d1O O CA\u00cdDA.",
    "MANGAS SUELTAS.": "ASEGURAR LA MANGA AL POSTE EN CONFIGURACI\u00d3N TIPO 'FIGURA 8', CONFORME AL EST\u00c1NDAR.",
    "MANGAS ABIERTAS/DA\u00d1ADAS.": "REEMPLAZAR TAPAS Y SELLOS, GARANTIZANDO EL CIERRE HERM\u00c9TICO Y LA PROTECCI\u00d3N DEL EMPALME CONTRA AGENTES EXTERNOS.",
    "RESERVAS SUELTAS.": "REORGANIZAR Y ASEGURAR LA RESERVA EN 'FIGURA 8' CONFORME A LO ESTABLECIDO.",
    "CRUCES DE V\u00cdAS BAJOS.": "AJUSTAR LA ALTURA DEL CABLE ELEV\u00c1NDOLO A LA DISTANCIA REGLAMENTARIA O REPORTAR PARA LA IMPLEMENTACI\u00d3N DE UNA SOLUCI\u00d3N ESTRUCTURAL.",
    "VEGETACI\u00d3N SOBRE FIBRA/MANGA.": "REALIZAR LA PODA O RETIRO DE VEGETACI\u00d3N QUE COMPROMETA LA INTEGRIDAD O SEGURIDAD DEL CABLE. EN CASO DE REQUERIR PERMISOS, DOCUMENTAR LA NOVEDAD.",
    "LOCALIZACI\u00d3N DE MANGA.": "DOCUMENTAR LA UBICACI\u00d3N MEDIANTE COORDENADAS GPS Y REGISTRO FOTOGR\u00c1FICO PARA ACTUALIZACI\u00d3N DE INVENTARIO.",
    "DOCUMENTACI\u00d3N UNIFILAR DE HILOS.": "DOCUMENTAR O SOLICITAR LA PROGRAMACI\u00d3N DE TRABAJO PARA OBTENER LA INFORMACI\u00d3N; UTILIZAR UN SEGUIDOR DE SE\u00d1AL.",
    "L\u00cdNEA EL\u00c9CTRICA EN MAL ESTADO.": "DOCUMENTAR EL RIESGO Y SOLICITAR AL COORDINADOR EL REPORTE AL \u00c1REA DE REGULATORIO.",
    "REGENERACI\u00d3N URBANA.": "ESTABLECER CONTACTO CON EL CONSORCIO, DOCUMENTAR LA AFECTACI\u00d3N Y COORDINAR, JUNTO CON EL COORDINADOR DE FO, LAS MEDIDAS DE MITIGACI\u00d3N, ASIGNANDO TAREAS A LOS DEPARTAMENTOS COMPETENTES COMO REGULATORIO, OBRA CIVIL Y \u00c1REAS INVOLUCRADAS.",
    "AMPLIACI\u00d3N DE V\u00cdA.": "DOCUMENTAR, REGISTRAR EL CONTACTO DEL RESPONSABLE DE LA OBRA Y COORDINAR MEDIDAS DE MITIGACI\u00d3N DE LA INFRAESTRUCTURA CON EL COORDINADOR DE FO.",
    "CABLE LASTIMADO.": "DOCUMENTAR E INFORMAR PARA PROGRAMAR EL CAMBIO DEL TRAMO DE CABLE.",
    "FIBRA INSTALADA INCORRECTAMENTE SOBRE MORDAZA.": "CORREGIR LA INSTALACI\u00d3N SEPARANDO ADECUADAMENTE EL CABLE DE FIBRA DEL MENSAJERO CONFORME A LA NORMATIVA T\u00c9CNICA.",
    "POZO SIN TAPA O EN MAL ESTADO.": "SOLICITAR LA EJECUCI\u00d3N DE TRABAJOS DE OBRA CIVIL PARA SU INSTALACI\u00d3N O CORRECCI\u00d3N.",
    "REPINTADO DE POZO.": "REALIZAR EL PINTADO DEL POZO TELCONET CON EL C\u00d3DIGO ASIGNADO POR GIS.",
    "REPINTADO DE POSTE.": "REALIZAR EL PINTADO DEL POSTE TELCONET CON EL C\u00d3DIGO ASIGNADO POR GIS.",
    "ELEMENTOS SIN ETIQUETAS ACR\u00cdLICAS.": "VERIFICAR, COLOCAR ETIQUETA ACR\u00cdLICA Y ETIQUETAR CON EL C\u00d3DIGO DE RUTA.",
    "RIESGO DE DERRUMBE O DESLAVE.": "DOCUMENTAR EL RIESGO Y SOLICITAR AL COORDINADOR LA REUBICACI\u00d3N DEL RECORRIDO DEL CABLE.",
    "RIESGO DE INUNDACIONES.": "DOCUMENTAR EL RIESGO Y SOLICITAR AL COORDINADOR LA REUBICACI\u00d3N DEL RECORRIDO DEL CABLE.",
    "RIESGO DE INCENDIO.": "DOCUMENTAR EL RIESGO Y SOLICITAR AL COORDINADOR LA REUBICACI\u00d3N DEL RECORRIDO DEL CABLE.",
}

HERR = ["Cintur\u00f3n y Linea de Vida","Casco","Escalera de 24 pies","Escalera de 28 pies","Escalera de 32 pies","Conos reflectivos","Caja para herramientas","Juego de destornilladores","Martillo mediano","Estiletes","Cortafr\u00edo","Alicate","Llave francesa","Juego de rachet","Pares de guantes aislantes","Tecle","Machete","Cizalla","Pata de cabra","Flejadora (M\u00e1quina Eriband)","Extensi\u00f3n con foco","Motosierra","Tijeras met\u00e1licas","Arco de sierra","Binoculares","Parasol","Remolque / Carrete para F.O."]
EQUI = ["Fusionadora","Cortadora de fibra","Bobina de lanzamiento","OTDR con cargador","Llave Acsys","GPS","Inversor","Etiquetadora"]
MATE = ["Fibra 48h (500mt)","Mangas de 48h y/o 144h (2 m\u00ednimo)","Rollo de cinta Eriband 3/4\"","Hebillas para cinta Eriband 3/4\"","Hojas de sierra","Patchcord de fibra","Adaptadores (Simplex-Duplex)","Paquetes de amarras","Mesas pl\u00e1sticas","Sillas pl\u00e1sticas","Cuchillos","Poleas","Sogas de nylon medianas","Sogas de nylon gruesas","Repelente contra insectos","Repelente contra abejas y avispas"]

LOGO_B64 = "iVBORw0KGgoAAAANSUhEUgAAAqcAAACoCAYAAAGqK3duAAAAAXNSR0IArs4c6QAAAARnQU1BAACxjwv8YQUAAAAJcEhZcwAAIdUAACHVAQSctJ0AACTYSURBVHhe7d0NkBxnfefxljDGYCAYCBxg7CPxG7IlWZqZlS3JXmu1uzMt2ZYsTffaWLLedmZn15J2pxdXpY4iDrnUcUnq6iAXwhXhOJJwBK54uQS4I7niOIKPCpfwfgFycEBMyhAOcPDZBozMXj+zz7P7TM9/ZrpnemZndr+fql9pp/vpp3ta7b//2pnpcQAASNWBxaWWQUzqZI1M1eeqfcuxl0VPsAks2xYvrTs59glUJ/TGw61PKFowJyl6Uu3H9sksVH9Xb4mWoifWPqn2CU2d2YkRfWxIy82y5eWb68Y0XbflYmH9MnuZvTyq3XqbdFJ7djKN6AGax3bs5ca2iUvrHtuy3o/EdfYc5s9c8TO1n3Pe4yvrpdikZa3YJ1b9XAjeodf0SPQAzWM70vKc/0TDMhNpvIlh/6xEx0Vjk5a1o8b39OocQOXS7Mf0j+rnv545OXO1fujMTM+8Uf/oVMpztRNj/ozNnFD150Y7uT1jTuauexr/K+Akd8GcWOmkcmK70Oykokv2ieWEpqRQ/fjKSUWKOKEAAGwUheof6p/QFdXgm+SDv697jJgKwZO1E2YafJPoyykq9gk2QYR9cqInsNVrVCZowj5J9kls9hqVCmIwJ8ucxO2H1KsVjSc0dWaHNvU4579JP1omjVNWl6++sJeduuDsPJIX1y0/Xt1OJeM/WbfM/LztrpfUfo6yx7Vjn1hzlbpBD0+oIh2gWWavs382pGVGs3VqWc7/St1683O7ZYa0rJXoSe3ZyTSkAzTL2qXV2LjrjFzxf4rr7RjRx3EUggu1k7qeRV+0sx9PT09fo6IfimM7etHPrZ61fv5Z7Wd0SZ1MUwrs3Dq95IyfP6BHITHppKqo5c6Dm/UoJCJdreqxG7xBj0BHoie1UP2pXoOu7CuvXqVIkTqpAAAAALB+ucFfrPwSxERSOPf8luuBro1YF5kd+xdJJtI4k0L1S3pGoAvSxaUiXZB21Mt1JtJ6FWleKYXqb+qjAdqQLiDp4jNpdZHm5xvnigboSvSCki5ElWYXanR7OwMjetDtkmQbaWwc0W2i2XbsUj1ylXr/lTTWzhZv+SYONmmcRBoTXdYsScaadMK+wKLzXT3WeJHa403c4O16tgFnPxEVib0++ibBKHusSjP2mO0HX6GXtmdv1440NrrMfmyWKdJy+3GzNyo2Y2+rkiZz0dmfxFcX6avzq4/ti3MomSdiIrHX9+JCbTUuyt4m6zd/65k9TsWQlhn2umii6wfpQjXURbj7+PL8ppoO9cVps0+eiiQ6plnijlV3uYmSxqm0tkncRqWZOGMUe5wZG13WLJI4Y9KS9R519k6/Xz/CMGr1BmLzBmOV48ePX6EXr6iUZh/SP9axtysWi8/Wi1eo5erP0vTMT0ql2VHzuCdMBbVj/wcyWqpf5waf0lsCa8C+GFXsi7VZotsAfVFYeO3KRZdfkC9OE/sCVa94AX23cgEKF6iKWe8GX9VbAGtmk3ixmosUGCj2xcoFCgAAAAAAAAAAgHS5i5/VPwEDxryGb9KMG3y6tl76fBmQukL1WMPFacetfkNcbgL0jHTBSV+1a96A0ixA6szF5VblC9KO+Vz/tjvk9Sr2Bdsu6ssSAJE7f5940ahIF54dc6FK61RuK8vzSgHaki4cFenis9PuQpXmjAZIRLqIVKQLUEXdfKLVhSrNFQ3QEeliki5ClW4uUqBr6ovJoheWdDE2u1Cj29oBUuVWv113gUUvRhXpQt17qv7CtDNQth68rO7A20UZ8beL6xrzzuXx1rI47PHNIpHGRRMljTHf1hplj6k99t9at6xZclPLX/UorWuWbRONdytsJx+8re5Ci86pLtJr9tcvs8ebDA37iUiiF2o7ccfa41biPxKu2eRki1MN64zocpWs98/CNZucnPfH4npDWmcSFV0XvVCT6mbbZiarIysXnD2/SrSa2heniuc9Q88yJOwnI+nFhWqPaTUuKsl20tjosqz35brHGf+x2jjFXl57PIAXqnLVuWetXHxm/myx/kK1L1CVoWSejIokzv/6bc2W2+wxOz1fL23P3i6O6Hj7sVmmSMsbHsf4X38rccd1qnah3r08/3WTqxfq0F+gRrsT2IuKmim+rm7cVYVn6TWt2duotCKNlZYZGf9zDetNlEGtqLbaxRrOry7SnUeWnFtOLl+gu08/T48YYu1OYDf/mGoWQ1rXLDl/+b5IGe8ecX2zvCzznNp2SnSdJDrGjEv6j6koe0wv7bmvvpquG+1OYC8vVGPk6KQ4LuO/TY+Q5fwvituN+G/WI+pFxzWTnfpMw7hhuVAVdaFieLW7gW759NytzcZMT09fo9apP/WiFTOl2de2mtte1+4YumJ6UftdTmbZyjpuKTnw2l0kar2JXlRHWl4oFJ7VartyufzMcmn2vfb66TOVT9ZW9kL0wtwf7tOu5tH1wJqJXoz2hdosZizQN4XFRxNfqCr2NipAz0UvOunCtBMdD/SNfeFJF6cdeyzQd+biU79flS5QFS5SrLmx6itWLkLpIlXhIsVAKARPNL1QuUgxUNTF6AbyhQoMlGhV5SLFwLLvpAIMrPzC+PKFymv+AAAAAAAAAAAAAAAAAAAAAAAAa2Zi8VW1T8s15oweAQBY5T2jdpNcuXDKacYNPi+OjwYAhlq++i2xuDWLW11ydgv3IJPGxk0+uFMfDQAMODd4s1jIVAqqQB5vLJBJo+70pOaaPL/kTJxdvpv+vsqSc1t5ydl7MtzHMXk76ZjSCgB0ZN+5XxaLih1V2KSiliTqS52iyfny2FZRxVc6xm7jOEP27bwABkOcIhqN6hilAhc30WJ69Zg8rlnU3fWk4+omAJAqqdA0y74Oi2pGfy2+HWlcs0jH0kny1TfqZw0APSQVICnjZ+Wi1yrRYnrtuDwumlvPyMeQJDdVn62fIQD0mVSUolEvIEkFsFmiBVUaE4203zhxF5/SzwQABoRUrOzkF+RCKMUupjcekseYdPKiUz54mz5q1Nl68DLxJKcVY8TfLq7vPO/UM4dzi+uX0ysZ7won539V3GfW/wcnV7xXj0xHphjU5pX2l/MedrLFzt4fKM1nJzv1Az0yPmkeE2PEf6u4vtPkpq7RM7d/Tt1k28Slei+9U6h+WCxiJtJx2Xl1vr6gSmNMpPmbhX/Kx0BBbW7n0VvEObtNxn9S76Fexv++OL7bZPz36j00ksa3yhXh9dKOtJ2JQUFtb7JaEAub6iilYzOxi+mOI/IYad5o8sEFfSRIlfQXYhJXq4Ka89+kR3VGmtMksdGLxHnsJHXl6CX6p0YZ713iPkwy/gf0yHhy/iPiPCa5o9frkcukMXGz/dAL9Cz1pLEmRquCuu2ul+hRvSHt02RQSQVPOn4Vu6BG191yunEeO/ngD/Qe0TPRvxQ7caXVoUqkcSZJXHniEnEOlWzxg3pUenL+t8R9qaRBmlcl663+RyOtN7FJ6+287Pbn6JHx5kyrQ+2ENI/JoIsWwOjx33i4eTGNbmuCPov+xdiJaxgKqiLNYZK2jFcV96OS8f5Wj+pMzntKnFfFJq03aUYaGzcGBbU7phiqj5Haxy91p5PzFNGBYv+FRRPX0PyTP5T1fl+cqy5tCl6mOCFvFyZKGhNN1nutHi0b8f6juJ2drP+YHr1KGmcSh7Rdqxj8kz8dpkCa448WU7O+sPDnegusOfOXIyWutXpRKkly/hN6xnoZ/4I4vpNcd/hFelbZjuJV4nad5vKbWr8yK21jkpQ0RzRGL1+Uakfa3mRYucE3a8eviul1+eXOddux/r2ABgAbQiH4Pyud6krHGuzQawGkpVKeW1LRDxPxPO9itW15uvJA+Xj5CvVz6UxlWq9uq1KafSjJvksnS1tr+zg9e7BcLu9XP588efJqvbotNd6OWlaarrxf/Vw+M+upP0+G+6gNHhZu9ccNxVKK6l7VLfpU5yqtj8ZxNi/vAEBsdoGxo1c31Wxc3O2VJAW12/3NTM8Va2OPVczvZzerx+XS3Bf040THPnCkohiN/auNONl7Sp7HrX5Y7xUA1qnJ+VmxAJok+ehqnDS7kcrE2VfpIwKAIScVOZM9J+TimDTtPiAAAOuKVOhUpAKZJK1u8QcA61Y+eFQsfFKhjJOxSuNcKuq+AwCwIUhFUCqYrbJvpnEON/jveg8AsIFMBm+pK4b7Z+XCKWUsHGtv6waP6FkBYAOzC6NUPKOZOFdfTAEAlsmzI7GK6vgchRQAYmlVVCfOU0gBIJkHNzd8C0BhYcnZ4l2sBwAAErn5+LdrxXTPyTv0EgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABlE+OOm4iw87o3PP1UsAALEUqj9c+VI9OwCAFgrnLheLpx03+KoeHU8+uLN+juB7eg0ArCNu9d/VF7s2GV/I6i3rHVg8HhbapxrGNyR4n94CAIbZg5vlIhc31X3y8pgBgKGVn79NLGytMlap/558E2ls3ADA0ClUvyQWtGj23y8XTSl77pPniBMAGAq3l58jFjGT/bNygUwSN1gKi/SSM3l+ORNnw3nnlpzbykvO3pPLxXb3scbt1DjpmNLILQsv02cAADrkBvNigdl1T2NBS5qr9jVGGtcuu483Hl9aGT9/hT4TAJCQVFSikYpakmS99IqpdHzdJr/wt/psAEACUkFplVtPy4Utbm483FhItx+Sx7aKdGzdBgA6JhWVdpGKW9xcO95YTKVxrSIdUzfJL1yvzwYAdKEQPCEWmWbZNyMXuTiJFtKkxTQ/Lx9TJ8kv/F99BgAgJbE+YWRFKnRxEi2k107I46Ts7uLtU9EAQM/kg78SC4+UZm+8b5doMc0clcdJkY4jaXaffp5+tgDQQ27wR2IRkiIVvHaJFlNpjBRp/0nCq/QA+s5dqIgFKRr1Jnup8LVKJ8VU2neSAMCaUa9wS4UpGqn4Ncv1B+oLqXosjbPTzSecCuf36mcDAGtovHyFWKTsTJyTi6CU6NuipDF2Ov2Ekxt8Xz8DABggUsGyIxVCKXYhjVNMpX21CwAMrjb3KY37u9MkxVTaT6vwxvsWpBOcVnL+J/Ve0t+PkSlOiOtV1Lpeyfi/Gz6/Rxv2mfV/5mSLn9ej0rI53NcXnezUhcb9eT92Roof1eOS2eXtbZgvmh3+r+vR8WSLo+I8KjuP3qJH9fB68L8hrk8r/SAVMRPpmKKxC+mOu+QxKqo4S/uQUqh+Sx8dmpJOclpZL8X0xru2i/MnScb/93q29nJTXxPnSJKM/1t6tubiFFOT7YdeoLdqjWKaDqmgmUjHZccuptJ6lbFZeW4piEk60Wll2Iup6jKlebuNKjhRV45eIo5NI80kKaYqOe8pvWVzFNP0SIVNRToukx1H2hfTm++V541G3d0fKcpNvUX8C1HZ4l2sR7UnbW/SjV4U06z3Z+J8UnLeyXCLzWE21f7Met8Ux9nZdser1G5W5Pyfi+OkqN+rLe9rU6KOOev/pLYvW9JiatJKGsW016R9qmT9OT1icEhFrtXvTrfevlpIr5uUx0hzRoMe2GjFNOc/Is5lslw8k9l15Gr9UyNpH3aWC2c8W4ovFOewY2tVTHNTPxWXm2T8x/Qs9Sim6ZOK3cjdjcevcs3+1WKq7mkaXS/NZQc9tJGKaebIHnEek7RlWvwaIevdr0cll/M+K86pol6wMlp2pv5NjjN6kbzOShTFtDfyi082FD7pOZhCqhJd1+oFJ7fJ1z8jRf0opnGT9b6gZ1uVZjGV5jDpBWk/Jt2S5jQx2hZTLev/QB6jk/P/UY9Mp5jGTe7wi/RsyUhzqQxyMVXcxU/XFcBdQnfarJjuK9cXT5PCIrfH6xuK6XJ6QdqPStb7kB7ROWleEyNuMa2J2aVSTHurEPxFXTGMPgdTSG+4vX65vY0J+mwjFdOdrf55XDyuR6VH2o9Jt6Q5TYxExVRr92JZ1v8NcbkKxTQdbvXUSkHc9ZrV48/5cldqF1CV8Wrju0nQBxvtBShpHpOc/3d6VDquC4uBtB+TbXfdoEfGl5l6nTiXia2TYrpss7xNm/A70/QUAm+lOJrjV/csjRZTu4jyWfo1ttGKqSLNFc3OlsUmPG/ew+J2Gf9tesSyHUez4jg7uan/p0c3seVi8dNR0Vx+07P1Bss6L6bL2r3zIRqKaboK555fK5J7Tywf//UH64up+o58U0gxAAbpn/l2jFbFNGlsGe9D4pi04jgXLe9Iy3lPi+PSSE54j6nSbTE1xO2FpP3PfDtJSNurDFsxVQ7OXlYrlur4TSFVHepoSXeji2N6JNbcRi2mRs7zxbGdJM5n9nNTfydu20kyxT/Vs8rSKqaK+n22OI8Vimnv7D1pdaV3Lzn54Am9BgNjoxfTOt4zxO1aZYv3XL1xctlDrxTnbJUk+0uzmC7bJM+lQzHtLVNMAQA9YH5vauJWv6HXAEhDuTy3t1KeW6qUKkf1osRq2+voRS2dPn36edPT09eomO3MYz2kpaT7M44dO3bpyrbTlc/oxc6hQydeYJbPlCvDVWQKwaONhTL4vDP6YP3v46NjoskvvEePBNCJbotpuO0Ftb36eaY0+3Pzc1y1fSfYxh6fdNvydGVBjX/QeXCz+rNUqrhqufp55szsPyadb2BIxTEa9aq++jVG3PuX5oM79ewA4uimmFamZ79pClA0ekhbScbPlCqP2/swCYv46r0GWjDj1c+l0tzr9bbW/wwq3zM/D5X9s78kFsRoor8fNlFv+s/Py9uYuNWtem8AJCvFNBLPm2v5Ilb5zOwfq3FzkXGjo6MXqeVhFyi/FSvC7E8/bKo0Pfv3atyJEyf+iV60Qi0P1/+NfthUdF+lMzM/tB/PTM+dVI+np8/W3x5xGBSqfygWQTt79HtR40Z97t8VOtlC9Wm9VwBYh9RboaKFLxqpaHaSyfPWvAuPh3uPf5tIABh4duGUIhXGbiLtIx88pI8GAIaYVOBM9s/KRbGT3HpG3oeJW13URwQAQ0h9b5NU3EykwthJpLlN8sGX9dEAwBArLPy1WORUpMKYNK26Ujd4RB8FAKwDUqEzkQpkkkhzqriLsd7FAQDDRSp4KlKBjJvbZuQ5VQBg3ZKK3lgXL0RJ86kAwLqWXzgkFj+pULbLvkrjPCoAsCHkq481FMBd98gFs1Wic6gAwIYSLYL5BblgNov0VdEAsCFFi6FUNJsluq36UkUA2LDsgqi+6kQqnNGMn40WUgDY4AqLd9QVRql4RmOPv/LEJXomANjg1Pfnxy2mdlfqztduqA0AMEyBzJ+Xi6jJSiFdfKPeEgBQxxRKqYiqjN+vC271Y3oLAIBIFcvdx+ViqtYVql/SIwEATRWqc2J3qr58r7D4XT0KANCWu/BwQzHNV3+q1wIAYhurWIV0nveSAkDHTDEFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAgC7tPv0850DwH5wDi0tNk198QI8GAAAAUlCo/rrjBk+JzWe7bPEu1rMAAAAAMe2bv1ZsLruJG7xTz9650QcvCpvj0TDfEffRLpPzh/VMAAAAGCj5hesdt/p1sYlLO271cb3XeqMnLgmPw3fy1U+I26WVQvBdvUcAAACsqULwKbFh61XGZpecm16z5IxMrea2kjy2H1HvhQUAAECfZMrPdAoLvyI2Zt2mUF1yxueWnJvvrW82k0aau9dxF35PnyEAAACkzg2mai+PS41Y3LjBkjNWWXJ2H5ebyF4kPy8fS6+SD55wbqo+W5+1dO2dvSx8PtcuJ7izlkJw2slXH6jFDd5eS37hPeGfH9H5jM5XwzxSS6H6dO1Y3eqPw8d/xAfFAADAYBr1nhs2Lo81NFzNoprN/WflprCXyRSXnKv2tc/VYaTte5W1fPtANykEn9dXAAAAQJ9NVl8YNpUfEZuUdpEasn4n58uNqJTMUXmO1HO3fL4GOYXqr+orAgAAoA8Kwb9x3OrPxMakm4jNWR+T9eRGNJrr8vL2aWfinHyeBi3qWlBvBwAAAOgP7xliU5J23KrcpPUrNxyUm9Fo1Ev/0vZpZc998vkZpLjB98ILY9Py9QEAALAWJmdfKTYqaUa911Rq2PoRqRGN5voD8rZpRT1/6bwMQgrBb+krAQAAYMC4iz8QG5g0slYNqtSMRqPelypt220G9QNP4w+8XP+NAwAADAH18q7U1KSRXffIjVwvsvUOuRm1s/WgvG032XX3YP221A2+qf9mAQAAhphbfZfY7HQb1bxJTV3auXpMbkhXEq6Xtusm++fk59zvuNVf03+LAAAA60wh+LdiA9RNpMYu7YgNqZXth+TtOon6ClTpefYrheCCM1ndrv/GAAAANoCJ+V1iY9Rpot9pn2Z23CU3pHak7TrJWr2E7y5+zRk9cYn+2wEAANigJs9nxGapk/Tqq0mvm5AbUpM0PgQ1ugYfeFJfNQoAAABBofAspxA8KTZRSbLnhNz8dZNW7ze9dlzeJnb6+Q1P6mb4Czv1GQcAAEAsYmOVIGk2qOprSKWm1CTbxQ33x/rwgad89YfOFu9ifWYBAADQsXzwkNhwxcnu++SGMGnUTfWlplRliytv0zb3yMecVgoL/0qfQaADWw9eJl+4Q5Cc/0n9LFaN+NvFsYOTd+ojXZUpTgjj2kdtN6iuPHGJk/Ne72S9H4nHnmayxQ86u45crffcD5ucHUeOOhn/b8TjSTs57z85ucPX6H2nZ5e3V9xfp8n579YzpytbHBX31y47j96iZ1g14r9VHDsoyU01/j1n/G+IY4ch2yYu1c9i+BWqHxabsHbZe1I+N0lybYv3m6rv2pe2aZXJeflYu0mh+rSTn79Zny2gSzSn/c46a05HL3Ey3nfF4xuEZKY+oQ+0E5vDa+xPxHkHIVn/grPTK+ljTS7t5tRO1n+z3kv3aE5pTgfJgdceE5uzVhmdls9PnKiX7KWmVOWG2+VtmuWWU/LxdRo3eIRP16M3aE77neFuTq+983nisQxLslMXnB2HX62fTdTmsNH+X+J2w5JM8XX6ubTXy+bUTs57i95jZ2hOaU4HUX7hoNiwNcu+GfkctYu6d6nUmKpI45tFOqZOUgjep88AMOByU28R/2Nol168QbrT5jTnv0nP0H/D0Jzmil8XjyFJMv73xYYhKfW8M1NPivuIJuN/ynFGL9JbynL+b4vbJk326P3hbM9YnrRDNxZ/0cn67xbnT5qdR3fpWWX9ak7tZL3f1HuPbxCa0213vUTPMJyk59QuWX9Ob41WCud+0ckHF8RGLppOGtTr8nJjuu1OeXw042flY0mS8eqofrbAEKE57c4gN6e54kPivuMk48/oWQbTTu+EeNxxkvM+q2fpPdUY5fwnxOOIk2Y6bk79m5zskXF5XYJkvbI+ktZoTrsnPad2oTlNanPtfZdSc2dnLEGDqu5dKjWm1+yXx9vZ1cU3PLnBU/o5AUNsPTSnvUjW+4I+qtYGtTmV9tkuOe9pJ+NdoWcYXFnvD8Tjb5fc0c7f45mGrP8P4nG1y5biC/UMq7ppTm2ZI4fD/4n+XB4bM7mpX9GzNVqvL+vnDr9IH1XvSftvF5rTDj242XGrXxebPpOxWfmcR9PsW6F2hsul8SYTHfy2tDD/Dv0EgHWC5lTOMDenW7znivtslx1Hs3qGwdbJXQSyYWMzCKRja5ec/xW99aq0mlNbxt8jb5MgO44e07MtozntnrT/dqE57V6++pdiI6gycU4+73a23t7YmKqb8UtjVdStq6R9Ncv4+X7e5QToM5pTOcPdnF4s7rNdcv6n9QyDLet9WTz+dlG3kVpLGe8XxONql2zxQT3Dql40p7aMf6D2m3RxjpjJFIs0pymQ9t8uNKfpUff8lJrDQlU+9ybRxlRFGqcizR9NIfiuM3Lu+fqogHWO95x2ZxCbU2VH8Spxv3GS6+IWR/2S9X4sHnu7ZPzP6Rn6aMvF4Tl9XDyedsl5/0NPUq/Xzaktjd+oJg3vOV0lPad2oTlNXz64p6FhnDzf5Px7jY2p+n796LixSv180eQXP6j3DmwwNKfdGdTm1Mj4HxD3Hzfq9k2ZzDP1bGna7Gy96/Lwf6L/MjwXn6gl638s/LtcdK4cjXfvPfU2BOmYk2RncaueLV1Z75+L+4sb9T7QqwrNf0vSz+bUtrM4Vjs2ce4UQ3O6SnpO7UJz2jsHF29w3ODnKw2k1KBKt5BSH5Ay62++t7ERVVHzusHtek/ABrYemtOs96GwiTrS80jvyey4OT36BnEfaSRz9JA+ulXq2HvVVGSLXw6b4P/iqHtijkz9Tu3ntO/vmPX+LHwW8q2eMrc/J9zfd8Ttuk3G//byc5t6R/j4d5xc8X21uyCom+hL47tNznu/flatrVVzatvpZeR9pJA0m9Osf0r87yTtbBl9rj7SdEnPqV1oTnuvcO5ZK01lfr7+/KtP5NuNqX3DffV+1fqG9Cd6RgAreM9pkqR3E/5eRt26qKXRi8Im6D+L2w5iskd/Tx94PDs9X5xnUHPt3S/XRx7fIDSntszhDo+nSQb9PadSpJvwp0HaV7vQnPaP+n+hu/ijsMlcPf92Y6qS85ac3cdXG9LJ4APhlmv7HnhgoNGcJsk6aU6b2Fm8N9z2UXHOXkf9VjfnfdxxenBdKeoWWTmv+y8j6CbZqf+mj6Z7g9ac2kbuulzed4LQnK6S9tUuNKdrww2+1vB+0+13Ljn5BfWBpr16FIC2dhRvDv9jOp046j2DaVP3c5T2NSgZmWr8H+aOO18ujl3LZPz79NH1Rta/wclMeWFD+dvh/5D/NPzzK+F+V7+jP+d/y8kVvxSOea+TLT4QNmXu0Lzvb+uRXwqf3x1hI/v65Zfw1XPz/7f13H5SW5Yr/tfwOf9++I+Tmdon0tVXw/bTrjteWvd3HjfqW6z66cbiFvE42uWG8PlFZY/uFscOSkbu7c0nqaV9tUvmyHV6a6yFl2We47zqVvV+9njfOucufrb+pf6Fh51Ctf62bAAADLJyeW5vpTy3tJJS5ahe1Vczpdl5+zhmSpVv61WpKJdmP2bPHzd6866VSpW32/OWz8z+iV7VMzPTlb+s2+d05Ud6lagyXbnFHq8yU5rhpuy9VqgeWGkmC9UfOvmz43pNMvn5I3WNaZy4iw+HW/K2AADA4Fjr5vTMmTMvtfdfOjU7Xpmee8I8np6efaMemrpKafYhe996carCpvROM3/YHD59+nT5ZnufM6dminpoqkqlUs7ej53wHwL/Wg8zNoX/GHhUHlt5XI9BLxSC94lNY9y4i18JG9ozzt5zy69kFKrvF8d1Grf6a87oiXh3PQEAIA1r2ZyG+3vM7Dds4v5cL645c2r21MoxhSmfKqf+PrteNqfHjh17SdiM/mzl+Kdn695nHq77F6vrKk8fP3481Zvxl4TnFjasW+1lJ0+efOWZE+WCvezEoRMvUGNnyrMXzLLw+N6glqFHMuVnio1hJ7E/TCXlptcsObdOL9/4X9o+SfLV7zgHqvv0swAAIB3R5jRsooLZM7PbpFROV24IN+n6JcDSdH3jVC5XvtYsYZO00sCqFIvFZ+tputar5lT9tnFl3um5p6TnZTJTWm0Cw+2+F27e9XvwT5yY+acr+w9TOlXapVfVzExXPmGvr2W68ht6dU25NPM5sy5sTr+vF6OXJquvEZvApGn3jVNJM1pa/iCWanyl/TVLYfGjzu3Bi/WzAwAAwFDKB18Wm70kmTgrN5q9zq57lpxbTi85k5F7sEZTWHwybMZ/NXy2vN8VAABg4I3NvEJs6pJkdFpuIAclrRrYQvWnzt7Zy/TZAAAAwEBwqx8Vm7e42X1MbgwHIeqtAtIxxwkAAADWiPoNotSgxc2uu+XmcC3TzQey8sEX9ZkBAADAmilUf11s1uJEahDXKtLxxcnyPVkBAAAwUNzq42Lz1ir75+RGsd+Rjq1dCsGjjuPF+4YsAAAArIHx+18uNnKtsveE3DD2K9IxtYtHUwoAADA83OrHxaauWUbW6P2nSe+Ruu/cL+tnCAAAgKEyWn6x2OBJUZ+Ql5rHXsZN8OEn9VWsAAAAWAfyC+8RG75oxu+Xm8jUc7e8fynu4rv1swAAAMA6skls/qK59YzQTKYY9U1R0n6jyQd/pY8bAAAA61Y+uFNsBu3cdK/cWHYbdV9VaX923OrXw6Pkq0sBAAA2FLf6BbE5NJGay24j7cfEDX6kjwwAAAAbkvrAlBv8vEmzKDeYnUbah8noiUv0EQEAAGDDcxffJTaNqXxAqsV7TMcWt+gjAAAAACzj5V8QG8g9XdygX713VZpzYn5G7xUAAABo4cDiA3WNZO3l/Q5u0C81pm71XXovAAAAQAJu8IOVprKQ8Ab90cbUDT6jZwUAAAA6NHru8pUGc/+s3IhGY98uittCAQAAIHXu/EdqzeYtp+WGdCX6w0/cFgoAAAA9NXb+pWHT+ZRz02uEpjTMzceXG1MAAACgbwqLb2poTPfct+RMVrfrEQAAAECfXXP7i52sf1Q/AgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACBwnP8PFay+FiN6mMMAAAAASUVORK5CYII="

AZUL="0000FF"; GRIS="969696"; GRIS2="D9D9D9"; GRIS3="C0C0C0"
AZUL2="0070C0"; VERDE="00B050"; ROJO="FF0000"; BLANCO="FFFFFF"
CREMA="FFF2CC"; AZUL_LINK="0563C1"; GRIS_TXT="333333"

# ── TOTP ──────────────────────────────────────────────────────────────────────
def verificar_totp(codigo):
    if not TOTP_SECRET:
        return True
    try:
        secreto = base64.b32decode(TOTP_SECRET.upper().replace(" ",""), casefold=True)
        ahora = int(time.time()) // 30
        for delta in [0, -1, 1]:
            contador = struct.pack(">Q", ahora + delta)
            mac = hmac.new(secreto, contador, hashlib.sha1).digest()
            offset = mac[-1] & 0x0F
            c = struct.unpack(">I", mac[offset:offset+4])[0] & 0x7FFFFFFF
            if str(c % 1000000).zfill(6) == str(codigo).strip():
                return True
    except Exception:
        pass
    return False

# ── GEMINI ────────────────────────────────────────────────────────────────────
async def analizar_imagen(img_bytes):
    img_b64 = base64.b64encode(img_bytes).decode()
    prompt = "Analiza esta imagen de inspeccion de fibra optica. Si hay problema responde JSON: {\"tiene_novedad\": true, \"motivo\": \"NOMBRE EN MAYUSCULAS\", \"coordenadas\": \"\"}. Si todo bien: {\"tiene_novedad\": false, \"motivo\": \"\", \"coordenadas\": \"\"}. Solo JSON."
    payload = {"contents": [{"parts": [{"text": prompt}, {"inline_data": {"mime_type": "image/jpeg", "data": img_b64}}]}]}
    try:
        async with httpx.AsyncClient(timeout=30) as client:
            resp = await client.post(GEMINI_URL, json=payload)
            texto = resp.json()["candidates"][0]["content"]["parts"][0]["text"]
            texto = texto.replace("```json","").replace("```","").strip()
            r = json.loads(texto)
            if not r.get("tiene_novedad"):
                return None
            motivo = r.get("motivo","").upper()
            return {"motivo": motivo, "remedio": REMEDIOS.get(motivo, "DOCUMENTAR Y REPORTAR AL COORDINADOR."), "coordenadas": r.get("coordenadas",""), "tarea_pendiente": "", "foto_antes": img_bytes, "foto_despues": None}
    except Exception as e:
        logger.error("Gemini error: " + str(e))
        return None

# ── EXCEL ─────────────────────────────────────────────────────────────────────
def _logo_image():
    from openpyxl.drawing.image import Image as XLImage
    try:
        data = base64.b64decode(LOGO_B64)
        img = XLImage(io.BytesIO(data))
        img.width = 155
        img.height = 54
        return img
    except Exception as e:
        logger.warning("No se pudo cargar el logo: " + str(e))
        return None

def _insertar_foto_celda(ws, fila, col, img_bytes, ancho_px=255, alto_px=400):
    """Inserta una foto (bytes JPEG/PNG) ajustada dentro de la celda foto, con pequeño margen."""
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
    from openpyxl.drawing.xdr import XDRPositiveSize2D
    from openpyxl.utils.units import pixels_to_EMU
    if not img_bytes:
        return
    try:
        img = XLImage(io.BytesIO(img_bytes))
        img.width = ancho_px
        img.height = alto_px
        marker = AnchorMarker(col=col-1, colOff=pixels_to_EMU(4), row=fila-1, rowOff=pixels_to_EMU(4))
        size = XDRPositiveSize2D(pixels_to_EMU(ancho_px), pixels_to_EMU(alto_px))
        img.anchor = OneCellAnchor(_from=marker, ext=size)
        ws.add_image(img)
    except Exception as e:
        logger.warning("No se pudo insertar foto en celda: " + str(e))

def _insertar_logo_centrado(ws, fila, col_letra="A"):
    """Inserta el logo centrado dentro de la celda, con margen para que quede dentro del borde."""
    from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
    from openpyxl.drawing.xdr import XDRPositiveSize2D
    from openpyxl.utils.units import pixels_to_EMU
    img = _logo_image()
    if not img:
        return
    col_idx = ord(col_letra.upper()) - ord("A")
    marker = AnchorMarker(col=col_idx, colOff=pixels_to_EMU(14), row=fila-1, rowOff=pixels_to_EMU(8))
    size = XDRPositiveSize2D(pixels_to_EMU(img.width), pixels_to_EMU(img.height))
    img.anchor = OneCellAnchor(_from=marker, ext=size)
    ws.add_image(img)

HERR_FILAS = ["Cintur\u00f3n y Linea de Vida","Casco","Escalera de 24 pies","Escalera de 28 pies","Escalera de 32 pies","Conos reflectivos","Caja para herramientas","Juego de destornilladores","Martillo mediano","Estiletes","Cortafr\u00edo","Alicate","Llave francesa","Juego de rachet","Pares de guantes aislantes","Tecle","Machete","Cizalla","Pata de cabra","Flejadora (M\u00e1quina Eriband)","Extensi\u00f3n con foco","Motosierra","Tijeras met\u00e1licas","Arco de sierra","Binoculares","Parasol","Remolque / Carrete para F.O."]
EQUI_FILAS = ["Fusionadora","Cortadora de fibra","Bobina de lanzamiento","OTDR con cargador","Llave Acsys","GPS","Inversor","Etiquetadora"]
MATE_FILAS = ["Fibra 48h (500mt)","Mangas de 48h y/o 144h (2 m\u00ednimo)","Rollo de cinta Eriband 3/4\"","Hebillas para cinta Eriband 3/4\"","Hojas de sierra","Patchcord de fibra","Adaptadores (Simplex-Duplex)","Paquetes de amarras","Mesas pl\u00e1sticas","Sillas pl\u00e1sticas","Cuchillos","Poleas","Sogas de nylon medianas","Sogas de nylon gruesas","Repelente contra insectos","Repelente contra abejas y avispas"]

SOLUCIONES = {
    "HERRAJES EN MAL ESTADO.":"REALIZAR EL REEMPLAZO INMEDIATO DEL HERRAJE AFECTADO, GARANTIZANDO LA CORRECTA SUJECI\u00d3N DEL CABLE Y LA ESTABILIDAD MEC\u00c1NICA DEL TENDIDO.",
    "FALTA DE HERRAJES.":"INSTALAR LOS HERRAJES CONFORME A LA NORMATIVA T\u00c9CNICA, ASEGURANDO LA CORRECTA FIJACI\u00d3N DEL CABLE AL POSTE.",
    "POSTES EN MAL ESTADO.":"DOCUMENTAR MEDIANTE REGISTRO FOTOGR\u00c1FICO Y COORDENADAS, Y REPORTAR PARA GESTIONAR EL REEMPLAZO DEL POSTE CON LA ENTIDAD RESPONSABLE.",
    "POSTE(S) CAMBIADO(S).":"INSTALAR LOS HERRAJES NECESARIOS Y ASEGURAR CORRECTAMENTE EL CABLE AL NUEVO POSTE. DOCUMENTAR EL CAMBIO PARA ACTUALIZACI\u00d3N DE INVENTARIO.",
    "POSTES POR INSTALAR.":"DOCUMENTAR LA UBICACI\u00d3N EXACTA Y REPORTAR PARA LA COORDINACI\u00d3N E INSTALACI\u00d3N DEL NUEVO POSTE REQUERIDO.",
    "POSTE NUEVO INSTALADO \u2013 TN.":"DOCUMENTAR, ETIQUETAR CON C\u00d3DIGO DE IDENTIFICACI\u00d3N Y APLICAR PINTURA DE SE\u00d1ALIZACI\u00d3N CONFORME A EST\u00c1NDARES OPERATIVOS.",
    "POSTE NUEVO INSTALADO \u2013 EMPRESAS EL\u00c9CTRICAS.":"DOCUMENTAR, COLOCAR ETIQUETA ACR\u00cdLICA Y ASEGURAR EL CABLE DE FIBRA \u00d3PTICA CONFORME A LA NORMATIVA T\u00c9CNICA VIGENTE, RESPETANDO LAS DISTANCIAS DE SEGURIDAD EL\u00c9CTRICA.",
    "POSTES INCLINADOS.":"DOCUMENTAR MEDIANTE REGISTRO FOTOGR\u00c1FICO Y COORDENADAS, Y REPORTAR PARA GESTIONAR EL APLOME DEL POSTE CON EL CONTRATISTA.",
    "RETENIDA(S) EN MAL ESTADO.":"DOCUMENTAR MEDIANTE REGISTRO FOTOGR\u00c1FICO Y COORDENADAS, Y REPORTAR PARA GESTIONAR LA CORRECCI\u00d3N CON EL CONTRATISTA.",
    "RETENIDA(S) CORTADA(S).":"DOCUMENTAR MEDIANTE REGISTRO FOTOGR\u00c1FICO Y COORDENADAS, Y REPORTAR PARA GESTIONAR LA CORRECCI\u00d3N CON EL CONTRATISTA.",
    "VANOS POR RETEMPLAR.":"REALIZAR EL RETEMPLADO DEL CABLE PARA RESTABLECER LA TENSI\u00d3N ADECUADA Y EVITAR RIESGOS DE DA\u00d1O O CA\u00cdDA.",
    "MANGAS SUELTAS.":"ASEGURAR LA MANGA AL POSTE EN CONFIGURACI\u00d3N TIPO 'FIGURA 8', CONFORME AL EST\u00c1NDAR.",
    "MANGAS ABIERTAS/DA\u00d1ADAS.":"REEMPLAZAR TAPAS Y SELLOS, GARANTIZANDO EL CIERRE HERM\u00c9TICO Y LA PROTECCI\u00d3N DEL EMPALME CONTRA AGENTES EXTERNOS.",
    "RESERVAS SUELTAS.":"REORGANIZAR Y ASEGURAR LA RESERVA EN 'FIGURA 8' CONFORME A LO ESTABLECIDO.",
    "CRUCES DE V\u00cdAS BAJOS.":"AJUSTAR LA ALTURA DEL CABLE ELEV\u00c1NDOLO A LA DISTANCIA REGLAMENTARIA O REPORTAR PARA LA IMPLEMENTACI\u00d3N DE UNA SOLUCI\u00d3N ESTRUCTURAL.",
    "VEGETACI\u00d3N SOBRE FIBRA/MANGA.":"REALIZAR LA PODA O RETIRO DE VEGETACI\u00d3N QUE COMPROMETA LA INTEGRIDAD O SEGURIDAD DEL CABLE. EN CASO DE REQUERIR PERMISOS, DOCUMENTAR LA NOVEDAD.",
    "LOCALIZACI\u00d3N DE MANGA.":"DOCUMENTAR LA UBICACI\u00d3N MEDIANTE COORDENADAS GPS Y REGISTRO FOTOGR\u00c1FICO PARA ACTUALIZACI\u00d3N DE INVENTARIO.",
    "DOCUMENTACI\u00d3N UNIFILAR DE HILOS.":"DOCUMENTAR O SOLICITAR LA PROGRAMACI\u00d3N DE TRABAJO PARA OBTENER LA INFORMACI\u00d3N; UTILIZAR UN SEGUIDOR DE SE\u00d1AL.",
    "L\u00cdNEA EL\u00c9CTRICA EN MAL ESTADO.":"DOCUMENTAR EL RIESGO Y SOLICITAR AL COORDINADOR EL REPORTE AL \u00c1REA DE REGULATORIO.",
    "REGENERACI\u00d3N URBANA.":"ESTABLECER CONTACTO CON EL CONSORCIO, DOCUMENTAR LA AFECTACI\u00d3N Y COORDINAR, JUNTO CON EL COORDINADOR DE FO, LAS MEDIDAS DE MITIGACI\u00d3N, ASIGNANDO TAREAS A LOS DEPARTAMENTOS COMPETENTES COMO REGULATORIO, OBRA CIVIL Y \u00c1REAS INVOLUCRADAS.",
    "AMPLIACI\u00d3N DE V\u00cdA.":"DOCUMENTAR, REGISTRAR EL CONTACTO DEL RESPONSABLE DE LA OBRA Y COORDINAR MEDIDAS DE MITIGACI\u00d3N DE LA INFRAESTRUCTURA CON EL COORDINADOR DE FO.",
    "CABLE LASTIMADO.":"DOCUMENTAR E INFORMAR PARA PROGRAMAR EL CAMBIO DEL TRAMO DE CABLE.",
    "FIBRA INSTALADA INCORRECTAMENTE SOBRE MORDAZA.":"CORREGIR LA INSTALACI\u00d3N SEPARANDO ADECUADAMENTE EL CABLE DE FIBRA DEL MENSAJERO CONFORME A LA NORMATIVA T\u00c9CNICA.",
    "POZO SIN TAPA O EN MAL ESTADO.":"SOLICITAR LA EJECUCI\u00d3N DE TRABAJOS DE OBRA CIVIL PARA SU INSTALACI\u00d3N O CORRECCI\u00d3N.",
    "REPINTADO DE POZO.":"REALIZAR EL PINTADO DEL POZO TELCONET CON EL C\u00d3DIGO ASIGNADO POR GIS.",
    "REPINTADO DE POSTE.":"REALIZAR EL PINTADO DEL POSTE TELCONET CON EL C\u00d3DIGO ASIGNADO POR GIS.",
    "ELEMENTOS SIN ETIQUETAS ACR\u00cdLICAS.":"VERIFICAR, COLOCAR ETIQUETA ACR\u00cdLICA Y ETIQUETAR CON EL C\u00d3DIGO DE RUTA.",
    "RIESGO DE DERRUMBE O DESLAVE.":"DOCUMENTAR EL RIESGO Y SOLICITAR AL COORDINADOR LA REUBICACI\u00d3N DEL RECORRIDO DEL CABLE.",
    "RIESGO DE INUNDACIONES.":"DOCUMENTAR EL RIESGO Y SOLICITAR AL COORDINADOR LA REUBICACI\u00d3N DEL RECORRIDO DEL CABLE.",
    "RIESGO DE INCENDIO.":"DOCUMENTAR EL RIESGO Y SOLICITAR AL COORDINADOR LA REUBICACI\u00d3N DEL RECORRIDO DEL CABLE.",
    "NO SE REGISTRAN NOVEDADES DURANTE LA INSPECCI\u00d3N.":"NO SE ENCUENTRAN NOVEDADES QUE SIGNIFIQUEN RIESGOS EN EL CABLE DE LA RED INTERURBANO.",
}

# ── Bordes ─────────────────────────────────────────────────────────────────────
_THIN = Side(style="thin", color="FF000000")
_MED  = Side(style="thick", color="FF000000")
_BORDE = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_BORDE_GRUESO = Border(left=_MED, right=_MED, top=_MED, bottom=_MED)

def _marco_grueso(ws, r1, c1, r2, c2):
    """Aplica borde medium SOLO en el contorno exterior de un rango ya existente (sin tocar lo interior)."""
    for r in range(r1, r2+1):
        for c in range(c1, c2+1):
            cell = ws.cell(r, c)
            existing = cell.border
            left   = _MED if c == c1 else (existing.left if existing else _THIN)
            right  = _MED if c == c2 else (existing.right if existing else _THIN)
            top    = _MED if r == r1 else (existing.top if existing else _THIN)
            bottom = _MED if r == r2 else (existing.bottom if existing else _THIN)
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)

def _hdr(ws, fila, col_ini, col_fin, texto, fg=AZUL, color=BLANCO, bold=True, size=11, borde=None):
    from openpyxl.styles import Alignment
    if col_fin > col_ini:
        ws.merge_cells(start_row=fila, start_column=col_ini, end_row=fila, end_column=col_fin)
    b = borde if borde else _BORDE
    for cc in range(col_ini, col_fin+1):
        cell = ws.cell(fila, cc)
        cell.border = b
    c = ws.cell(fila, col_ini, texto)
    c.font = Font(bold=bold, name="Calibri", size=size, color=color)
    c.fill = PatternFill("solid", fgColor=fg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return c

def _lbl(ws, fila, col, texto, bg=GRIS2, bold=True):
    from openpyxl.styles import Alignment
    c = ws.cell(fila, col, texto)
    c.font = Font(bold=bold, name="Calibri", size=11, color="000000")
    if bg: c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.border = _BORDE
    return c

def _val(ws, fila, col_ini, col_fin, texto, bold=False, borde=None):
    from openpyxl.styles import Alignment
    if col_fin > col_ini:
        ws.merge_cells(start_row=fila, start_column=col_ini, end_row=fila, end_column=col_fin)
    b = borde if borde else _BORDE
    for cc in range(col_ini, col_fin+1):
        cell = ws.cell(fila, cc)
        cell.border = b
    c = ws.cell(fila, col_ini, texto)
    c.font = Font(bold=bold, name="Calibri", size=11, color="000000")
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    return c

def generar_excel(datos):
    r   = datos["recorrido"]
    ciu = datos["ciu"]
    nch = datos["mpriu"].get("novedades_check", {})
    wb  = Workbook()

    # ══ HOJA 1: REPORTES_DE_RECORRIDOS — hasta 20 novedades ════════════
    ws1 = wb.active
    ws1.title = "REPORTES_DE_RECORRIDOS"
    ws1.column_dimensions["A"].width = 41.14
    ws1.column_dimensions["B"].width = 35.57
    ws1.column_dimensions["C"].width = 32.29
    ws1.column_dimensions["D"].width = 30.57
    ws1.row_dimensions[2].height = 60
    ws1.cell(2,1).border = _BORDE
    _insertar_logo_centrado(ws1, 2, "A")
    _hdr(ws1, 2, 2, 3, "REPORTE DE RECORRIDOS DE MANTENIMIENTO PREVENTIVO PARA RUTAS INTERURBANAS", fg="FFFFFF", color="000000")
    _val(ws1, 2, 4, 4, "Código: FOR FO 02\nVersión: 3 (28/05/2021)", bold=True)
    _marco_grueso(ws1, 2, 1, 2, 4)
    ws1.row_dimensions[4].height = 24
    _hdr(ws1, 4, 1, 4, "REPORTE DE RECORRIDO DE RUTAS INTERURBANAS DE F. O.")
    _marco_grueso(ws1, 4, 1, 4, 4)
    ws1.row_dimensions[5].height = 38; ws1.row_dimensions[6].height = 38
    ws1.merge_cells(start_row=5,start_column=1,end_row=6,end_column=1)
    for fr in (5,6): ws1.cell(fr,1).border = _BORDE
    _lbl(ws1, 5, 1, "FECHA Y HORA DEL RECORRIDO", bg=GRIS)
    ws1.cell(5,1).alignment = ws1.cell(5,1).alignment.copy(horizontal="center")
    _lbl(ws1, 5, 2, "FECHA"); _lbl(ws1, 5, 3, "HORA INICIO"); _lbl(ws1, 5, 4, "HORA FIN")
    _val(ws1, 6, 2, 2, r.get("fecha","")); _val(ws1, 6, 3, 3, r.get("hora_inicio","")); _val(ws1, 6, 4, 4, r.get("hora_fin",""))
    _marco_grueso(ws1, 5, 1, 6, 4)
    for i,(label,valor) in enumerate([("NOMBRE DE LA RUTA",r.get("nombre_ruta","")),("CÓDIGO DE CUADRILLA",r.get("codigo_cuadrilla","")),("NODO INICIAL",r.get("nodo_inicial",""))]):
        f=7+i; ws1.row_dimensions[f].height=38
        _lbl(ws1, f, 1, label); _val(ws1, f, 2, 4, valor, bold=True)

    # Hasta 20 novedades — cada bloque ocupa 6 filas desde fila 10
    fila = 10
    novedades = r.get("novedades", [])[:20]
    for nov in novedades:
        num=str(nov.get("numero",""))
        fila_inicio_bloque = fila
        ws1.row_dimensions[fila].height=38
        ws1.merge_cells(start_row=fila,start_column=1,end_row=fila+1,end_column=1)
        for fr in (fila, fila+1): ws1.cell(fr,1).border = _BORDE
        _lbl(ws1, fila, 1, "FECHA Y HORA NOVEDAD # "+num, bg=GRIS)
        ws1.cell(fila,1).alignment = ws1.cell(fila,1).alignment.copy(horizontal="center")
        _lbl(ws1, fila, 2, "FECHA"); _lbl(ws1, fila, 3, "HORA INICIO"); _lbl(ws1, fila, 4, "HORA FIN")
        fila+=1; ws1.row_dimensions[fila].height=38
        _val(ws1, fila, 2, 2, nov.get("fecha","")); _val(ws1, fila, 3, 3, nov.get("hora_inicio","")); _val(ws1, fila, 4, 4, nov.get("hora_fin",""))
        fila+=1
        for label,key in [("MOTIVO APARENTE DE LA NOVEDAD","motivo"),("REMEDIO DEFINITIVO A LA NOVEDAD","remedio"),("TAREA PENDIENTE (por regulatorio/obra civil, contratista)","tarea_pendiente"),("COORDENADAS SITIO DE LA NOVEDAD (Grados decimales)","coordenadas")]:
            ws1.row_dimensions[fila].height=42
            _lbl(ws1, fila, 1, label, bg=GRIS3); _val(ws1, fila, 2, 4, nov.get(key,"")); fila+=1
        _marco_grueso(ws1, fila_inicio_bloque, 1, fila_inicio_bloque+1, 4)
        _marco_grueso(ws1, fila_inicio_bloque+2, 1, fila-1, 4)

    # Si hay menos de 20, dejar bloques vacios numerados igual que el original
    for n_extra in range(len(novedades)+1, 21):
        fila_inicio_bloque = fila
        ws1.row_dimensions[fila].height=38
        ws1.merge_cells(start_row=fila,start_column=1,end_row=fila+1,end_column=1)
        for fr in (fila, fila+1): ws1.cell(fr,1).border = _BORDE
        _lbl(ws1, fila, 1, "FECHA Y HORA NOVEDAD # "+str(n_extra), bg=GRIS)
        ws1.cell(fila,1).alignment = ws1.cell(fila,1).alignment.copy(horizontal="center")
        _lbl(ws1, fila, 2, "FECHA"); _lbl(ws1, fila, 3, "HORA INICIO"); _lbl(ws1, fila, 4, "HORA FIN")
        fila+=1; ws1.row_dimensions[fila].height=38
        _val(ws1, fila, 2, 2, ""); _val(ws1, fila, 3, 3, ""); _val(ws1, fila, 4, 4, "")
        fila+=1
        for label in ["MOTIVO APARENTE DE LA NOVEDAD","REMEDIO DEFINITIVO A LA NOVEDAD","TAREA PENDIENTE (por regulatorio/obra civil, contratista)","COORDENADAS SITIO DE LA NOVEDAD (Grados decimales)"]:
            ws1.row_dimensions[fila].height=42
            _lbl(ws1, fila, 1, label, bg=GRIS3); _val(ws1, fila, 2, 4, ""); fila+=1
        _marco_grueso(ws1, fila_inicio_bloque, 1, fila_inicio_bloque+1, 4)
        _marco_grueso(ws1, fila_inicio_bloque+2, 1, fila-1, 4)

    for label,valor in [("NODO FINAL",r.get("nodo_final","")),("LIDER DE CUADRILLA QUE ELABORA INFORME",r.get("lider","")),("AYUDANTE TÉCNICO",r.get("ayudante","")),("COORDINADOR FIBRA ÓPTICA",r.get("coordinador",""))]:
        ws1.row_dimensions[fila].height=38
        _lbl(ws1, fila, 1, label); _val(ws1, fila, 2, 4, valor)
        fila+=1

    fila_ini_fotos = fila
    fila_fotos_val = fila
    _lbl(ws1, fila, 1, "FOTOS ANEXAS AL REPORTE (INDIQUE CUANTAS)", bg=CREMA)
    ws1.cell(fila,1).font = Font(bold=True, underline="single", name="Calibri", size=11, color=AZUL_LINK)
    _fotos_txt = str(r.get("fotos_total",0)) if r.get("fotos_total",0) else ""
    _val(ws1, fila, 2, 4, _fotos_txt)
    ws1.cell(fila,2).fill = PatternFill("solid", fgColor=CREMA)
    ws1.cell(fila,2).font = Font(bold=False, name="Calibri", size=11, color=GRIS_TXT)
    ws1.row_dimensions[fila].height=38
    fila+=1
    _lbl(ws1, fila, 1, "OBSERVACIONES GENERALES", bg=CREMA)
    ws1.cell(fila,1).font = Font(bold=True, name="Calibri", size=11, color=GRIS_TXT)
    _val(ws1, fila, 2, 4, r.get("observaciones",""))
    ws1.cell(fila,2).fill = PatternFill("solid", fgColor=CREMA)
    ws1.cell(fila,2).font = Font(bold=False, name="Calibri", size=11, color=GRIS_TXT)
    ws1.row_dimensions[fila].height=38
    fila+=1
    _marco_grueso(ws1, fila_ini_fotos, 1, fila-1, 4)

    # ══ HOJA 2: FOTOS_ANEXAS_AL_REPORTE — hasta 20 novedades ══════════
    ws2=wb.create_sheet("FOTOS_ANEXAS_AL_REPORTE")
    ws2.column_dimensions["A"].width=2.71
    ws2.column_dimensions["B"].width=19.71; ws2.column_dimensions["C"].width=58.71; ws2.column_dimensions["D"].width=23.71; ws2.column_dimensions["E"].width=35.71
    ws2.column_dimensions["F"].width=10.14
    ws2.row_dimensions[1].height=4
    ws2.row_dimensions[2].height=60
    ws2.row_dimensions[3].height=4
    ws2.cell(2,2).border = _BORDE_GRUESO
    _insertar_logo_centrado(ws2, 2, "A")
    _hdr(ws2, 2, 3, 4, "REPORTE DE RECORRIDOS DE MANTENIMIENTO PREVENTIVO PARA RUTAS INTERURBANAS", fg="FFFFFF", color="000000", borde=_BORDE_GRUESO)
    _val(ws2, 2, 5, 5, "Código: FOR FO 02\nVersión: 3 (28/05/2021)", bold=True, borde=_BORDE_GRUESO)
    ws2.row_dimensions[4].height = 18
    _hdr(ws2, 4, 2, 5, "FOTOS DE LAS ACCIONES CORRECTIVAS", borde=_BORDE_GRUESO)
    ws2.row_dimensions[5].height = 4
    ws2.row_dimensions[6].height = 4
    for fr in (5, 6):
        ws2.cell(fr,2).border = Border(left=_THIN)
        ws2.cell(fr,5).border = Border(right=_THIN)

    # Bloque NODO INICIO — etiqueta B fusionada en altura con foto (filas 7-8)
    from openpyxl.styles import Alignment as _Al
    ws2.merge_cells(start_row=7, start_column=2, end_row=8, end_column=2)
    for fr in (7,8): ws2.cell(fr,2).border = _BORDE_GRUESO
    c_lbl = ws2.cell(7,2,"NODO INICIO RECORRIDO")
    c_lbl.font = Font(bold=True, name="Calibri", size=11, color="000000")
    c_lbl.alignment = _Al(horizontal="center", vertical="center", wrap_text=True)
    _hdr(ws2, 7, 3, 3, "FOTO", borde=_BORDE_GRUESO)
    _hdr(ws2, 7, 4, 5, "NOMBRE DEL NODO", borde=_BORDE_GRUESO)
    ws2.row_dimensions[8].height = 315
    _val(ws2, 8, 3, 3, "", borde=_BORDE_GRUESO)
    _val(ws2, 8, 4, 5, r.get("nodo_inicial",""), bold=True, borde=_BORDE_GRUESO)
    _insertar_foto_celda(ws2, 8, 3, r.get("nodo_inicial_foto"))

    f2=10
    for n_idx in range(1, 21):
        nov = novedades[n_idx-1] if n_idx <= len(novedades) else None
        ws2.merge_cells(start_row=f2, start_column=2, end_row=f2+1, end_column=2)
        for fr in (f2, f2+1): ws2.cell(fr,2).border = _BORDE_GRUESO
        c_lbl = ws2.cell(f2,2,"NOVEDAD # "+str(n_idx))
        c_lbl.font = Font(bold=True, name="Calibri", size=11, color="000000")
        c_lbl.alignment = _Al(horizontal="center", vertical="center", wrap_text=True)
        _hdr(ws2, f2, 3, 3, "ANTES DEL MANTENIMIENTO", borde=_BORDE_GRUESO)
        _hdr(ws2, f2, 4, 5, "DESPU\u00c9S DEL MANTENIMIENTO", borde=_BORDE_GRUESO)
        f2+=1; ws2.row_dimensions[f2].height=315
        _val(ws2, f2, 3, 3, "", borde=_BORDE_GRUESO); _val(ws2, f2, 4, 5, "", borde=_BORDE_GRUESO)
        if nov:
            _insertar_foto_celda(ws2, f2, 3, nov.get("foto_antes"))
            _insertar_foto_celda(ws2, f2, 4, nov.get("foto_despues"))
        f2+=2

    ws2.merge_cells(start_row=f2, start_column=2, end_row=f2+1, end_column=2)
    for fr in (f2, f2+1): ws2.cell(fr,2).border = _BORDE_GRUESO
    c_lbl = ws2.cell(f2,2,"NODO FINAL DEL RECORRIDO")
    c_lbl.font = Font(bold=True, name="Calibri", size=11, color="000000")
    c_lbl.alignment = _Al(horizontal="center", vertical="center", wrap_text=True)
    _hdr(ws2, f2, 3, 3, "FOTO", borde=_BORDE_GRUESO)
    _hdr(ws2, f2, 4, 5, "NOMBRE DEL NODO", borde=_BORDE_GRUESO)
    f2 += 1
    ws2.row_dimensions[f2].height = 315
    _val(ws2, f2, 3, 3, "", borde=_BORDE_GRUESO)
    _val(ws2, f2, 4, 5, r.get("nodo_final",""), bold=True, borde=_BORDE_GRUESO)
    _insertar_foto_celda(ws2, f2, 3, r.get("nodo_final_foto"))
    _marco_grueso(ws2, f2-1, 2, f2, 5)

    if GENERAR_HOJAS_EXTRA:
        # ══ HOJA 3: MANGAS — hasta 40 mangas (20 pares) ════════════════════
        ws3=wb.create_sheet("MANGAS")
        ws3.column_dimensions["A"].width=20
        ws3.column_dimensions["B"].width=18; ws3.column_dimensions["C"].width=28; ws3.column_dimensions["D"].width=18; ws3.column_dimensions["E"].width=28
        ws3.row_dimensions[2].height=60
        ws3.cell(2,1).border = _BORDE
        _insertar_logo_centrado(ws3, 2, "A")
        _hdr(ws3, 2, 3, 4, "REPORTE DE RECORRIDOS DE MANTENIMIENTO PREVENTIVO PARA RUTAS INTERURBANAS", fg="FFFFFF", color="000000")
        _val(ws3, 2, 5, 5, "C\u00f3digo: FOR FO 02\nVersi\u00f3n: 3 (28/05/2021)", bold=True)
        ws3.row_dimensions[4].height = 30
        _hdr(ws3, 4, 2, 5, "FOTOS DE LAS MANGAS  DESDE EL NODO A AL B", fg="00133A")
        mangas=datos.get("mangas",[])[:40]
        f3=6
        n_pares = max(20, (len(mangas)+1)//2)
        for i in range(0, n_pares*2, 2):
            m1 = mangas[i]   if i   < len(mangas) else {}
            m2 = mangas[i+1] if i+1 < len(mangas) else {}
            _hdr(ws3, f3, 2, 2, "NOMBRE:", fg="1F4E79", size=10)
            _val(ws3, f3, 3, 3, m1.get("nombre",""))
            _hdr(ws3, f3, 4, 4, "NOMBRE:", fg="1F4E79", size=10)
            _val(ws3, f3, 5, 5, m2.get("nombre","")); f3+=1
            ws3.row_dimensions[f3].height = 315
            _val(ws3, f3, 3, 3, "")
            _val(ws3, f3, 5, 5, ""); f3+=1
            for label,k in [("DERIVACI\u00d3N:","derivacion"),("COORDENADAS:","coordenadas"),("OBSERVACI\u00d3N:","observacion")]:
                _hdr(ws3, f3, 2, 2, label, fg="1F4E79", size=10)
                _val(ws3, f3, 3, 3, m1.get(k,""))
                _hdr(ws3, f3, 4, 4, label, fg="1F4E79", size=10)
                _val(ws3, f3, 5, 5, m2.get(k,"")); f3+=1
            f3 += 1

        # ══ HOJA 4: INVENTARIO DE HILOS — hasta 6 nodos x 48 hilos ════════
        ws4=wb.create_sheet("INVENTARIO DE HILOS EN NODO")
        ws4.column_dimensions["A"].width=20
        ws4.column_dimensions["B"].width=8; ws4.column_dimensions["C"].width=30
        ws4.column_dimensions["D"].width=4; ws4.column_dimensions["E"].width=10; ws4.column_dimensions["F"].width=10
        ws4.row_dimensions[2].height=60
        ws4.cell(2,1).border = _BORDE
        _insertar_logo_centrado(ws4, 2, "A")
        _hdr(ws4, 2, 3, 6, "REPORTE DE RECORRIDOS DE MANTENIMIENTO PREVENTIVO PARA RUTAS INTERURBANAS", fg="FFFFFF", color="000000")
        _val(ws4, 2, 7, 7, "Código: FOR FO 02\nVersión: 3 (28/05/2021)", bold=True)

        hilos = datos["hilos"].get("filas", [])
        # Si no hay hilos, igual mostramos el nodo principal vacio (1 nodo)
        if not hilos:
            _lbl(ws4, 4, 1, "NODO: ", bg=None); _val(ws4, 4, 3, 3, r.get("nodo_final",""))
            _lbl(ws4, 5, 1, "NOMBRE ODF DE RUTA:", bg=None); _val(ws4, 5, 3, 3, datos["hilos"].get("posicion_odf",""))
            _lbl(ws4, 6, 1, "TIPO DE FIBRA", bg=None); _val(ws4, 6, 3, 3, "48H 4 BUFFER")
            _lbl(ws4, 7, 1, "CODIGO COLOR", bg=None); _val(ws4, 7, 3, 3, "INTERNACIONAL")
            for col,txt in [(1,"PAR"),(2,"HILO"),(3,"NOMENCLATURA"),(5,"RACK #")]:
                _val(ws4, 8, col, col, txt, bold=True)
            _val(ws4, 9, 2, 2, "SIN CAMBIOS EN ODF EN ESTE RECORRIDO")
        else:
            # Particionar hilos en bloques de hasta 48 (1 nodo cada uno), hasta 6 nodos
            BLOQUE = 48
            bloques = [hilos[i:i+BLOQUE] for i in range(0, len(hilos), BLOQUE)] or [[]]
            bloques = bloques[:6]
            fila_n = 4
            for bi, bloque in enumerate(bloques):
                _lbl(ws4, fila_n, 1, "NODO: ", bg=None); _val(ws4, fila_n, 3, 3, r.get("nodo_final",""))
                _lbl(ws4, fila_n+1, 1, "NOMBRE ODF DE RUTA:", bg=None); _val(ws4, fila_n+1, 3, 3, datos["hilos"].get("posicion_odf",""))
                _lbl(ws4, fila_n+2, 1, "TIPO DE FIBRA", bg=None); _val(ws4, fila_n+2, 3, 3, "48H 4 BUFFER")
                _lbl(ws4, fila_n+3, 1, "CODIGO COLOR", bg=None); _val(ws4, fila_n+3, 3, 3, "INTERNACIONAL")
                for col,txt in [(1,"PAR"),(2,"HILO"),(3,"NOMENCLATURA"),(5,"RACK #")]:
                    _val(ws4, fila_n+4, col, col, txt, bold=True)
                fh = fila_n + 5
                for idx,h in enumerate(bloque):
                    par = idx // 2 + 1
                    if idx % 2 == 0:
                        _val(ws4, fh, 1, 1, str(par))
                    _val(ws4, fh, 2, 2, str(idx+1))
                    _val(ws4, fh, 3, 3, h.get("descripcion",""))
                    fh+=1
                fila_n = fh + 2

        # ══ HOJA 5: Checklist CIU — SIN LOGO ════════════════════════════════
        ws5=wb.create_sheet("Checklist CIU")
        for col,w in [("A",9),("B",26),("C",11),("D",21),("E",11),("F",14),("G",11),("H",14)]: ws5.column_dimensions[col].width=w
        ws5.row_dimensions[2].height=46
        _hdr(ws5, 2, 2, 7, "CHECKLIST CUADRILLA INTERURBANA")
        _val(ws5, 2, 8, 8, "C\u00f3digo: FOR FO 05\nVersi\u00f3n: 3 (26/06/2025)", bold=True)
        _lbl(ws5, 4, 2, "Fecha del Recorrido", bg=None); _val(ws5, 4, 3, 4, r.get("fecha",""))
        _lbl(ws5, 4, 5, "Hora Inicio", bg=None); _val(ws5, 4, 6, 6, r.get("hora_inicio",""))
        _lbl(ws5, 4, 7, "Hora Fin", bg=None); _val(ws5, 4, 8, 8, r.get("hora_fin",""))
        f5=5
        for label,valor in [("Nombre de Ruta",r.get("nombre_ruta","")),("Nodo Inicio",r.get("nodo_inicial","")),("Nodo Final",r.get("nodo_final","")),("Distancia de la Ruta",ciu.get("distancia_ruta","")),("Lider de Cuadrilla",r.get("lider","")),("Veh\u00edculo Placa",ciu.get("vehiculo_placa","")),("Coordinador Fibra \u00d3ptica",r.get("coordinador",""))]:
            _lbl(ws5, f5, 2, label, bg=None); _val(ws5, f5, 3, 8, valor); f5+=1
        f5+=1
        for sec,items,data_s in [("HERRAMIENTAS Y EPP",HERR_FILAS,ciu.get("herramientas",{})),("EQUIPOS ELECTRONICOS",EQUI_FILAS,ciu.get("equipos",{})),("MATERIALES E INSUMOS",MATE_FILAS,ciu.get("materiales",{}))]:
            _hdr(ws5, f5, 2, 3, sec, fg=AZUL2)
            _hdr(ws5, f5, 4, 4, "CANTIDAD", fg=AZUL2)
            _hdr(ws5, f5, 5, 8, "OBSERVACIONES", fg=AZUL2); f5+=1
            for nombre in items:
                info=data_s.get(nombre,{})
                cant=info.get("cantidad",0) if isinstance(info,dict) else int(info or 0)
                obs=info.get("obs","NINGUNA") if isinstance(info,dict) else ("BUEN ESTADO" if cant>0 else "NINGUNA")
                _val(ws5, f5, 2, 2, nombre)
                ws5.cell(f5,4,cant).alignment = ws5.cell(f5,2).alignment.copy(horizontal="center")
                ws5.cell(f5,4).border = _BORDE
                bg_o = VERDE if obs=="BUEN ESTADO" else (ROJO if obs=="MAL ESTADO" else "808080")
                _hdr(ws5, f5, 5, 8, obs, fg=bg_o, bold=False); f5+=1

        # ══ HOJA 6: Checklists MPRIU — SIN LOGO ════════════════════════════
        ws6=wb.create_sheet("Checklists MPRIU")
        for col,w in [("A",9),("B",26),("C",11),("D",32),("E",11),("F",14),("G",11),("H",14)]: ws6.column_dimensions[col].width=w
        ws6.row_dimensions[2].height=46
        _hdr(ws6, 2, 2, 7, "CHECKLIST DE RECORRIDO DE MANTENIMIENTO PREVENTIVO DE RUTAS INTERURBANAS")
        _val(ws6, 2, 8, 8, "C\u00f3digo: FOR FO 08\nVersi\u00f3n: 02 (28/05/2021)", bold=True)
        _lbl(ws6, 4, 2, "Fecha del Recorrido", bg=None); _val(ws6, 4, 3, 4, r.get("fecha",""))
        _lbl(ws6, 4, 5, "Hora Inicio", bg=None); _val(ws6, 4, 6, 6, r.get("hora_inicio",""))
        _lbl(ws6, 4, 7, "Hora  Fin", bg=None); _val(ws6, 4, 8, 8, r.get("hora_fin",""))
        f6=5
        for label,valor in [("Nombre de Ruta",r.get("nombre_ruta","")),("Nodo Inicio",r.get("nodo_inicial","")),("Nodo Final",r.get("nodo_final","")),("Distancia de la Ruta",ciu.get("distancia_ruta","")),("Lider de Cuadrilla",r.get("lider","")),("Veh\u00edculo Placa",ciu.get("vehiculo_placa","")),("Coordinador Fibra \u00d3ptica",r.get("coordinador",""))]:
            _lbl(ws6, f6, 2, label, bg=None); _val(ws6, f6, 3, 8, valor); f6+=1
        f6+=1
        ws6.row_dimensions[f6].height=30
        _hdr(ws6, f6, 2, 2, "NOVEDAD", fg=AZUL2)
        _hdr(ws6, f6, 3, 3, "CHECK", fg=AZUL2)
        _hdr(ws6, f6, 4, 7, "SOLUCI\u00d3N", fg=AZUL2)
        _hdr(ws6, f6, 8, 8, "CANTIDAD", fg=AZUL2); f6+=1
        for novedad in NOVEDADES_MPRIU:
            ws6.row_dimensions[f6].height=43
            info=nch.get(novedad,{}); tiene=info.get("check",False); cant=info.get("cantidad",0); chk="SI" if tiene else "NO"
            sol=SOLUCIONES.get(novedad,"DOCUMENTAR Y REPORTAR AL COORDINADOR.")
            _val(ws6, f6, 2, 2, novedad)
            _hdr(ws6, f6, 3, 3, chk, fg=(VERDE if tiene else ROJO))
            _val(ws6, f6, 4, 7, sol)
            ws6.cell(f6,8,cant if tiene else 0).alignment = ws6.cell(f6,2).alignment.copy(horizontal="center")
            ws6.cell(f6,8).border = _BORDE
            f6+=1
        ws6.row_dimensions[f6].height=60
        _lbl(ws6, f6, 2, "Observaciones:", bg=None); _val(ws6, f6, 3, 8, r.get("observaciones",""))

    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf.read()

def datos_vacios():
    return {
        "recorrido":{"fecha":"","hora_inicio":"","hora_fin":"","nombre_ruta":"","codigo_cuadrilla":"","nodo_inicial":"","nodo_final":"","lider":"","ayudante":"","coordinador":"","fotos_total":0,"observaciones":"","novedades":[]},
        "ciu":{"vehiculo_placa":"","distancia_ruta":"","herramientas":{},"equipos":{},"materiales":{}},
        "mpriu":{"novedades_check":{},"observaciones":""},
        "mangas":[],"hilos":{"posicion_odf":"","filas":[]},
    }

def novedad_vacia(numero):
    ahora=datetime.now()
    return {"numero":numero,"fecha":ahora.strftime("%d/%m/%Y"),"hora_inicio":ahora.strftime("%H:%M:%S"),"hora_fin":ahora.strftime("%H:%M:%S"),"motivo":"","remedio":"","tarea_pendiente":"","coordenadas":"","foto_antes":None,"foto_despues":None}

def nombre_archivo(datos):
    ruta=datos["recorrido"]["nombre_ruta"].split()[0].replace("/","-") if datos["recorrido"]["nombre_ruta"] else "RUTA"
    return "FOR_FO_02_"+ruta+"_"+datetime.now().strftime("%Y%m%d_%H%M")+".xlsx"

# ── AUTENTICACION ─────────────────────────────────────────────────────────────
async def start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id in USUARIOS_AUTENTICADOS:
        return await menu_principal(update, ctx)
    await update.message.reply_text(
        "RecorridosIA - Acceso restringido" + chr(10) + chr(10) +
        "Ingresa tu correo y codigo de 6 digitos:" + chr(10) + chr(10) +
        "email: tucorreo@telconet.ec" + chr(10) + "totp: 123456",
        reply_markup=ReplyKeyboardRemove()
    )
    return ESPERANDO_TOTP

async def handler_totp(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    texto=update.message.text.strip().lower(); email=""; codigo=""
    for linea in texto.splitlines():
        if linea.startswith("email:"): email=linea.replace("email:","").strip()
        elif linea.startswith("totp:"): codigo=linea.replace("totp:","").strip()
    if not email or not codigo:
        await update.message.reply_text("Formato incorrecto. Usa:" + chr(10) + "email: tucorreo@telconet.ec" + chr(10) + "totp: 123456")
        return ESPERANDO_TOTP
    if not email.endswith(DOMINIO):
        await update.message.reply_text("Solo correos @"+DOMINIO)
        return ESPERANDO_TOTP
    if verificar_totp(codigo):
        USUARIOS_AUTENTICADOS.add(update.effective_user.id)
        nombre=email.split("@")[0].upper()
        await update.message.reply_text("Acceso autorizado. Bienvenido "+nombre)
        return await menu_principal(update, ctx)
    await update.message.reply_text("Codigo incorrecto." + chr(10) + "email: tucorreo@telconet.ec" + chr(10) + "totp: 123456")
    return ESPERANDO_TOTP

async def menu_principal(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    teclado=[["Generar Informe","Nueva Ruta Base"],["Mis Rutas","Ayuda"]]
    await update.message.reply_text("RecorridosIA - Menu principal",reply_markup=ReplyKeyboardMarkup(teclado,resize_keyboard=True))
    return MENU_PRINCIPAL

# ── MENU DE PESTANAS ──────────────────────────────────────────────────────────
async def generar_informe(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id not in USUARIOS_AUTENTICADOS:
        return await start(update, ctx)
    if "datos" not in ctx.user_data:
        ctx.user_data["datos"] = datos_vacios()
    return await tab_menu(update, ctx)

async def tab_menu(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id not in USUARIOS_AUTENTICADOS:
        return await start(update, ctx)
    if "datos" not in ctx.user_data:
        ctx.user_data["datos"] = datos_vacios()
    datos=ctx.user_data["datos"]; r=datos["recorrido"]
    ciu_ok=bool(datos["ciu"].get("vehiculo_placa"))
    mpriu_ok=bool(datos["mpriu"].get("novedades_check"))
    rep_ok=bool(r.get("nombre_ruta"))
    novedades=len(r.get("novedades",[]))
    man_ok=bool(datos.get("mangas"))
    hil_ok=bool(datos["hilos"].get("filas"))
    def v(ok): return "✅" if ok else "⬜"
    botones=[]
    if GENERAR_HOJAS_EXTRA:
        botones.append([InlineKeyboardButton(v(ciu_ok)+" Checklist CIU",callback_data="tab_1")])
        botones.append([InlineKeyboardButton(v(mpriu_ok)+" Checklists MPRIU",callback_data="tab_2")])
    botones.append([InlineKeyboardButton(v(rep_ok)+" REPORTES_DE_RECORRIDOS",callback_data="tab_reportes")])
    botones.append([InlineKeyboardButton(("✅" if novedades>0 else "⬜")+" FOTOS_ANEXAS_AL_REPORTE ["+str(novedades)+" nov]",callback_data="tab_fotos")])
    if GENERAR_HOJAS_EXTRA:
        botones.append([InlineKeyboardButton(v(man_ok)+" Mangas",callback_data="tab_5"),InlineKeyboardButton(v(hil_ok)+" Hilos ODF",callback_data="tab_6")])
    botones.append([InlineKeyboardButton("GENERAR EXCEL",callback_data="tab_generar")])
    teclado=InlineKeyboardMarkup(botones)
    total_pestanas = 4 if GENERAR_HOJAS_EXTRA else 2
    completadas=sum([ciu_ok,mpriu_ok,rep_ok,novedades>0]) if GENERAR_HOJAS_EXTRA else sum([rep_ok,novedades>0])
    msg="INFORME FOR FO 02 — Completado: "+str(completadas)+"/"+str(total_pestanas)+chr(10)+"Selecciona la pestana:"
    if update.callback_query:
        await update.callback_query.answer()
        try: await update.callback_query.edit_message_text(msg,reply_markup=teclado)
        except: await update.callback_query.message.reply_text(msg,reply_markup=teclado)
    else:
        await update.message.reply_text(msg,reply_markup=teclado)
    return TAB_MENU

async def tab_callback(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query=update.callback_query; await query.answer(); data=query.data
    volver=InlineKeyboardMarkup([[InlineKeyboardButton("Volver al menu",callback_data="tab_menu")]])

    if data=="tab_generar":
        return await enviar_excel(update, ctx)

    elif data=="tab_menu":
        return await tab_menu(update, ctx)

    elif data=="tab_1":
        msg=("CHECKLIST CIU - HERRAMIENTAS Y EPP"+chr(10)+"━━━━━━━━━━━━━━━━━━━━━━━━━━━━"+chr(10)+chr(10)+
             "Escribe cantidades separadas por coma:"+chr(10)+chr(10)+
             " 1. Cinturon y Linea de Vida"+chr(10)+" 2. Casco"+chr(10)+" 3. Escalera 24 pies"+chr(10)+
             " 4. Escalera 28 pies"+chr(10)+" 5. Escalera 32 pies"+chr(10)+" 6. Conos reflectivos"+chr(10)+
             " 7. Caja para herramientas"+chr(10)+" 8. Juego destornilladores"+chr(10)+" 9. Martillo mediano"+chr(10)+
             "10. Estiletes"+chr(10)+"11. Cortafrio"+chr(10)+"12. Alicate"+chr(10)+"13. Llave francesa"+chr(10)+
             "14. Juego de rachet"+chr(10)+"15. Guantes aislantes (pares)"+chr(10)+"16. Tecle"+chr(10)+
             "17. Machete"+chr(10)+"18. Cizalla"+chr(10)+"19. Pata de cabra"+chr(10)+"20. Flejadora Eriband"+chr(10)+
             "21. Extension con foco"+chr(10)+"22. Motosierra"+chr(10)+"23. Tijeras metalicas"+chr(10)+
             "24. Arco de sierra"+chr(10)+"25. Binoculares"+chr(10)+"26. Parasol"+chr(10)+"27. Remolque/Carrete FO"+chr(10)+chr(10)+
             "Ejemplo: 2,2,0,2,0,6,0,1,1,2,2,0,0,1,2,2,2,1,0,0,0,0,0,0,0,0,0")
        await query.edit_message_text(msg,reply_markup=volver)
        ctx.user_data["tab_actual"]="1"; return TAB_CIU_HERR

    elif data=="tab_2":
        msg="CHECKLISTS MPRIU"+chr(10)+"━━━━━━━━━━━━━━━━━━━━━━━━━━━━"+chr(10)+chr(10)+"Escribe los NUMEROS de las novedades (separados por coma):"+chr(10)+chr(10)
        for i,nov in enumerate(NOVEDADES_MPRIU,1): msg+=str(i).rjust(2)+". "+nov+chr(10)
        msg+=chr(10)+"━━━━━━━━━━━━━━━━━━━━━━━━━━━━"+chr(10)+"Ejemplo: 16,18"+chr(10)+"Sin novedades: 31"
        await query.edit_message_text(msg,reply_markup=volver)
        ctx.user_data["tab_actual"]="2"; return TAB_MPRIU

    elif data in ("tab_reportes","tab_fotos"):
        nombre_tab="REPORTES_DE_RECORRIDOS" if data=="tab_reportes" else "FOTOS_ANEXAS_AL_REPORTE"
        teclado2=InlineKeyboardMarkup([
            [InlineKeyboardButton("Manual (sin senial)",callback_data="rep_manual"),InlineKeyboardButton("Con IA (Gemini)",callback_data="rep_ia")],
            [InlineKeyboardButton("Volver al menu",callback_data="tab_menu")],
        ])
        await query.edit_message_text(nombre_tab+chr(10)+chr(10)+"Como quieres llenar esta pestana?",reply_markup=teclado2)
        return TAB_MENU

    elif data=="rep_manual":
        msg=("REPORTES - Modo Manual"+chr(10)+"━━━━━━━━━━━━━━━━━━━━━━━━━━━━"+chr(10)+chr(10)+
             "Ingresa los datos (uno por linea):"+chr(10)+chr(10)+
             "RUTA: GOSSEAL-MACHACHI   TAREA: 157415066"+chr(10)+"CUADRILLA: FO UIO INT 04"+chr(10)+
             "NODO_INI: GOSSEAL"+chr(10)+"NODO_FIN: MACHACHI"+chr(10)+"LIDER: RICHARD DAVID TAIPE COYAGO"+chr(10)+
             "AYUDANTE: JOSE LUIS ALLAICA CONDO"+chr(10)+"COORDINADOR: JUAN CARLOS YEPEZ ACAN"+chr(10)+
             "PLACA: PCO3940"+chr(10)+"DISTANCIA: 59KM"+chr(10)+"FECHA: HOY"+chr(10)+
             "HORA_INI: AHORA"+chr(10)+"HORA_FIN: AHORA"+chr(10)+"FOTOS: 6"+chr(10)+"OBS: texto o NINGUNA"+chr(10)+chr(10)+
             "Para novedades agrega:"+chr(10)+"NOV: VEGETACION SOBRE FIBRA/MANGA. | -0.477,-78.579"+chr(10)+chr(10)+
             "Cuando termines escribe: FIN")
        await query.edit_message_text(msg,reply_markup=volver)
        ctx.user_data["tab_actual"]="3"; ctx.user_data["novedades_manuales"]=[]
        return TAB_REPORTES

    elif data=="rep_ia":
        await query.edit_message_text(
            "REPORTES + FOTOS con Gemini IA"+chr(10)+"━━━━━━━━━━━━━━━━━━━━━━━━━━━━"+chr(10)+chr(10)+
            "Envia las fotos de la inspeccion."+chr(10)+"Gemini detectara automaticamente las novedades."+chr(10)+chr(10)+
            "Cuando termines escribe: LISTO",reply_markup=volver)
        ctx.user_data["tab_actual"]="4"; ctx.user_data["media_inspeccion"]=[]
        return TAB_NOVEDADES_IA

    elif data=="tab_5":
        ctx.user_data["datos"]["mangas"]=[]; ctx.user_data["manga_temp"]={}
        await query.edit_message_text(
            "MANGAS"+chr(10)+"━━━━━━━━━━━━━━━━━━━━━━━━━━━━"+chr(10)+chr(10)+
            "Voy a pedirte los datos de cada manga uno por uno."+chr(10)+chr(10)+
            "Nombre de la manga #1:"+chr(10)+"Ejemplo: UIO-B-MAC/GOS-F1-DER-01",reply_markup=volver)
        return MANGA_NOMBRE

    elif data=="tab_6":
        await query.edit_message_text(
            "INVENTARIO DE HILOS EN NODO"+chr(10)+"━━━━━━━━━━━━━━━━━━━━━━━━━━━━"+chr(10)+chr(10)+
            "Posicion del ODF:"+chr(10)+"Ejemplo: MAC-GOS-F01-R04-ODF02-48",reply_markup=volver)
        return HILO_ODF

    elif data in ("manga_der_si","manga_der_no"):
        derivacion="SI" if data=="manga_der_si" else "NO"
        ctx.user_data["manga_temp"]["derivacion"]=derivacion
        await query.edit_message_text(
            "Nombre: "+ctx.user_data["manga_temp"].get("nombre","")+chr(10)+
            "Derivacion: "+derivacion+chr(10)+chr(10)+
            "Coordenadas GPS de la manga:"+chr(10)+"Ejemplo: -0.477057,-78.579350")
        return MANGA_COORDS

    return TAB_MENU

# ── CIU HANDLERS ──────────────────────────────────────────────────────────────
async def tab_ciu_herr(update, ctx):
    valores=[v.strip() for v in update.message.text.replace(","," ").split()]
    herr={}; resumen=""
    for i,nombre in enumerate(HERR):
        cant=int(valores[i]) if i<len(valores) and valores[i].isdigit() else 0
        herr[nombre]={"cantidad":cant,"obs":"BUEN ESTADO" if cant>0 else "NINGUNA"}
        if cant>0: resumen+="  ✅ "+nombre+": "+str(cant)+chr(10)
    ctx.user_data["datos"]["ciu"]["herramientas"]=herr
    await update.message.reply_text(
        "✅ HERRAMIENTAS GUARDADAS"+chr(10)+"━━━━━━━━━━━━━━━━━━━━━━━━━━━━"+chr(10)+
        (resumen if resumen else "  Ninguna")+chr(10)+"━━━━━━━━━━━━━━━━━━━━━━━━━━━━"+chr(10)+chr(10)+
        "EQUIPOS ELECTRONICOS — escribe cantidades:"+chr(10)+chr(10)+
        " 1. Fusionadora"+chr(10)+" 2. Cortadora de fibra"+chr(10)+" 3. Bobina de lanzamiento"+chr(10)+
        " 4. OTDR con cargador"+chr(10)+" 5. Llave Acsys"+chr(10)+" 6. GPS"+chr(10)+
        " 7. Inversor"+chr(10)+" 8. Etiquetadora"+chr(10)+chr(10)+"Ejemplo: 1,2,0,1,1,0,1,1")
    return TAB_CIU_EQUI

async def tab_ciu_equi(update, ctx):
    valores=[v.strip() for v in update.message.text.replace(","," ").split()]
    equi={}; resumen=""
    for i,nombre in enumerate(EQUI):
        cant=int(valores[i]) if i<len(valores) and valores[i].isdigit() else 0
        equi[nombre]={"cantidad":cant,"obs":"BUEN ESTADO" if cant>0 else "NINGUNA"}
        if cant>0: resumen+="  ✅ "+nombre+": "+str(cant)+chr(10)
    ctx.user_data["datos"]["ciu"]["equipos"]=equi
    await update.message.reply_text(
        "✅ EQUIPOS GUARDADOS"+chr(10)+"━━━━━━━━━━━━━━━━━━━━━━━━━━━━"+chr(10)+
        (resumen if resumen else "  Ninguno")+chr(10)+"━━━━━━━━━━━━━━━━━━━━━━━━━━━━"+chr(10)+chr(10)+
        "MATERIALES E INSUMOS — escribe cantidades:"+chr(10)+chr(10)+
        " 1. Fibra 48h (500mt)"+chr(10)+" 2. Mangas 48h/144h"+chr(10)+" 3. Rollo cinta Eriband 3/4"+chr(10)+
        " 4. Hebillas Eriband 3/4"+chr(10)+" 5. Hojas de sierra"+chr(10)+" 6. Patchcord de fibra"+chr(10)+
        " 7. Adaptadores (Simplex-Duplex)"+chr(10)+" 8. Paquetes de amarras"+chr(10)+" 9. Mesas plasticas"+chr(10)+
        "10. Sillas plasticas"+chr(10)+"11. Cuchillos"+chr(10)+"12. Poleas"+chr(10)+
        "13. Sogas nylon medianas"+chr(10)+"14. Sogas nylon gruesas"+chr(10)+
        "15. Repelente insectos"+chr(10)+"16. Repelente abejas/avispas"+chr(10)+chr(10)+
        "Ejemplo: 335,2,1,6,0,2,10,2,0,0,0,1,1,0,0,0")
    return TAB_CIU_MATE

async def tab_ciu_mate(update, ctx):
    valores=[v.strip() for v in update.message.text.replace(","," ").split()]
    mate={}; resumen=""
    for i,nombre in enumerate(MATE):
        cant=int(valores[i]) if i<len(valores) and valores[i].isdigit() else 0
        mate[nombre]={"cantidad":cant,"obs":"BUEN ESTADO" if cant>0 else "NINGUNA"}
        if cant>0: resumen+="  ✅ "+nombre+": "+str(cant)+chr(10)
    ctx.user_data["datos"]["ciu"]["materiales"]=mate
    await update.message.reply_text(
        "✅ CHECKLIST CIU COMPLETO"+chr(10)+"━━━━━━━━━━━━━━━━━━━━━━━━━━━━"+chr(10)+
        (resumen if resumen else "  Ninguno")+chr(10)+"━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
    return await tab_menu(update, ctx)

# ── MPRIU HANDLERS ────────────────────────────────────────────────────────────
async def tab_mpriu(update, ctx):
    txt=update.message.text.strip().upper()
    if txt=="OK":
        await update.message.reply_text("✅ CHECKLIST MPRIU GUARDADO!")
        return await tab_menu(update, ctx)
    numeros=[n.strip() for n in txt.replace(","," ").split() if n.strip().isdigit()]
    nch={}
    for num_str in numeros:
        idx=int(num_str)-1
        if 0<=idx<len(NOVEDADES_MPRIU):
            nov=NOVEDADES_MPRIU[idx]
            nch[nov]={"check":True,"cantidad":1}
    ctx.user_data["datos"]["mpriu"]["novedades_check"]=nch
    cant=len([v for v in nch.values() if v.get("check")])
    resumen="".join(["  ✅ "+nov+chr(10) for nov in nch])
    await update.message.reply_text(
        "✅ "+str(cant)+" novedad(es) marcada(s):"+chr(10)+"━━━━━━━━━━━━━━━━━━━━━━━━━━━━"+chr(10)+
        (resumen if resumen else "  Sin novedades")+chr(10)+"━━━━━━━━━━━━━━━━━━━━━━━━━━━━"+chr(10)+chr(10)+
        "Ajusta cantidades en formato NUMERO:CANTIDAD"+chr(10)+"Ejemplo: 16:5,18:1"+chr(10)+"Si estan correctas escribe: OK")
    return TAB_MPRIU

# ── REPORTES HANDLERS ─────────────────────────────────────────────────────────
async def tab_reportes(update, ctx):
    lineas=update.message.text.strip().split(chr(10))
    datos=ctx.user_data["datos"]; r=datos["recorrido"]
    for linea in lineas:
        if ":" not in linea: continue
        clave,valor=linea.split(":",1); clave=clave.strip().upper(); valor=valor.strip()
        if valor.upper()=="HOY": valor=datetime.now().strftime("%d/%m/%Y")
        elif valor.upper()=="AHORA": valor=datetime.now().strftime("%H:%M:%S")
        if clave=="RUTA": r["nombre_ruta"]=valor.upper()
        elif clave=="CUADRILLA": r["codigo_cuadrilla"]=valor.upper()
        elif clave=="NODO_INI": r["nodo_inicial"]=valor.upper()
        elif clave=="NODO_FIN": r["nodo_final"]=valor.upper()
        elif clave=="LIDER": r["lider"]=valor.upper()
        elif clave=="AYUDANTE": r["ayudante"]=valor.upper()
        elif clave=="COORDINADOR": r["coordinador"]=valor.upper()
        elif clave=="PLACA": datos["ciu"]["vehiculo_placa"]=valor.upper()
        elif clave=="DISTANCIA": datos["ciu"]["distancia_ruta"]=valor.upper()
        elif clave=="FECHA": r["fecha"]=valor
        elif clave=="HORA_INI": r["hora_inicio"]=valor
        elif clave=="HORA_FIN": r["hora_fin"]=valor
        elif clave=="FOTOS": r["fotos_total"]=int(valor) if valor.isdigit() else 0
        elif clave=="OBS":
            if valor.upper()!="NINGUNA": r["observaciones"]=valor.upper()
    novedades_nuevas=[]
    for linea in lineas:
        if linea.upper().startswith("NOV:"):
            partes=linea[4:].strip().split("|")
            motivo=partes[0].strip().upper(); coords=partes[1].strip() if len(partes)>1 else ""
            remedio=REMEDIOS.get(motivo,"DOCUMENTAR Y REPORTAR AL COORDINADOR.")
            nov=novedad_vacia(len(novedades_nuevas)+1)
            nov["motivo"]=motivo; nov["remedio"]=remedio; nov["coordenadas"]=coords
            novedades_nuevas.append(nov)
            datos["mpriu"]["novedades_check"][motivo]={"check":True,"cantidad":1}
        elif linea.upper()=="FIN": break
    if novedades_nuevas: r["novedades"]=novedades_nuevas
    elif not r.get("novedades"):
        nov=novedad_vacia(1); nov["motivo"]=SIN_NOV_MOTIVO; nov["remedio"]=SIN_NOV_REMEDIO
        r["novedades"]=[nov]
    await update.message.reply_text(
        "✅ Datos guardados!"+chr(10)+"Ruta: "+r.get("nombre_ruta","")+chr(10)+
        "Novedades: "+str(len(r.get("novedades",[]))))
    return await tab_menu(update, ctx)

async def tab_novedades_ia(update, ctx):
    if "media_inspeccion" not in ctx.user_data: ctx.user_data["media_inspeccion"]=[]
    if update.message.photo:
        foto=await update.message.photo[-1].get_file()
        ctx.user_data["media_inspeccion"].append(bytes(await foto.download_as_bytearray()))
        await update.message.reply_text("Foto "+str(len(ctx.user_data["media_inspeccion"]))+" recibida. Envia mas o escribe LISTO")
        return TAB_NOVEDADES_IA
    if update.message.text and update.message.text.upper()=="LISTO":
        await update.message.reply_text("Analizando con Gemini IA...")
        media=ctx.user_data.get("media_inspeccion",[]); novedades=[]
        for img in media:
            r_ia=await analizar_imagen(img)
            if r_ia:
                n=novedad_vacia(len(novedades)+1); n.update(r_ia); novedades.append(n)
        if not novedades:
            n=novedad_vacia(1); n["motivo"]=SIN_NOV_MOTIVO; n["remedio"]=SIN_NOV_REMEDIO; novedades=[n]
        ctx.user_data["datos"]["recorrido"]["novedades"]=novedades
        ctx.user_data["datos"]["recorrido"]["fotos_total"]=len(media)
        for nov in novedades:
            m=nov["motivo"]
            if m!=SIN_NOV_MOTIVO:
                ctx.user_data["datos"]["mpriu"]["novedades_check"][m]={"check":True,"cantidad":ctx.user_data["datos"]["mpriu"]["novedades_check"].get(m,{}).get("cantidad",0)+1}
        await update.message.reply_text("✅ "+str(len(novedades))+" novedad(es) detectadas por la IA!")
        return await tab_menu(update, ctx)
    return TAB_NOVEDADES_IA

# ── MANGAS ────────────────────────────────────────────────────────────────────
async def recv_manga_nombre(update, ctx):
    txt=update.message.text.strip().upper()
    if txt=="FIN MANGAS":
        total=len(ctx.user_data["datos"]["mangas"])
        await update.message.reply_text("✅ "+str(total)+" manga(s) guardada(s).")
        return await tab_menu(update, ctx)
    ctx.user_data["manga_temp"]={"nombre":txt}
    teclado=InlineKeyboardMarkup([[InlineKeyboardButton("SI - Con derivacion",callback_data="manga_der_si"),InlineKeyboardButton("NO - Sin derivacion",callback_data="manga_der_no")]])
    await update.message.reply_text("✅ Nombre: "+txt+chr(10)+chr(10)+"Tiene derivacion esta manga?",reply_markup=teclado)
    return MANGA_COORDS

async def recv_manga_coords(update, ctx):
    ctx.user_data["manga_temp"]["coordenadas"]=update.message.text.strip()
    await update.message.reply_text("✅ Coordenadas: "+update.message.text.strip()+chr(10)+chr(10)+"Observacion de la manga:"+chr(10)+"Si no hay escribe: NINGUNA")
    return MANGA_OBS

async def recv_manga_obs(update, ctx):
    obs=update.message.text.strip()
    if obs.upper()=="NINGUNA": obs=""
    manga=ctx.user_data.pop("manga_temp",{})
    manga["observacion"]=obs; manga.setdefault("derivacion","NO")
    ctx.user_data["datos"]["mangas"].append(manga)
    total=len(ctx.user_data["datos"]["mangas"])
    teclado=InlineKeyboardMarkup([[InlineKeyboardButton("Terminar mangas",callback_data="tab_menu")]])
    await update.message.reply_text(
        "✅ Manga #"+str(total)+" guardada!"+chr(10)+chr(10)+
        "Nombre: "+manga.get("nombre","")+chr(10)+"Derivacion: "+manga.get("derivacion","NO")+chr(10)+
        "Coordenadas: "+manga.get("coordenadas","")+chr(10)+"Observacion: "+(obs or "NINGUNA")+chr(10)+chr(10)+
        "Nombre de la manga #"+str(total+1)+":"+chr(10)+"O escribe: FIN MANGAS",reply_markup=teclado)
    ctx.user_data["manga_temp"]={}
    return MANGA_NOMBRE

# ── HILOS ─────────────────────────────────────────────────────────────────────
async def recv_hilo_odf(update, ctx):
    ctx.user_data["datos"]["hilos"]["posicion_odf"]=update.message.text.upper()
    await update.message.reply_text("Ingresa hilos:"+chr(10)+"HILO, DESCRIPCION, ESTADO"+chr(10)+"Ejemplo: 1, TELCONET, OCUPADO"+chr(10)+"Cuando termines: FIN HILOS")
    return HILO_DATOS

async def recv_hilo_datos(update, ctx):
    if update.message.text.upper()=="FIN HILOS":
        await update.message.reply_text("✅ Hilos guardados!")
        return await tab_menu(update, ctx)
    partes=update.message.text.split(",")
    if len(partes)>=3:
        ctx.user_data["datos"]["hilos"]["filas"].append({"hilo_par":partes[0].strip(),"descripcion":partes[1].strip(),"estado":partes[2].strip().upper()})
    await update.message.reply_text("✅ Guardado. Siguiente o FIN HILOS:")
    return HILO_DATOS

# ── GENERAR EXCEL ─────────────────────────────────────────────────────────────
async def enviar_excel(update, ctx):
    try:
        msg=update.callback_query.message if update.callback_query else update.message
        if update.callback_query: await update.callback_query.answer()
        await msg.reply_text("Generando informe FOR FO 02...")
        if "datos" not in ctx.user_data: ctx.user_data["datos"]=datos_vacios()
        datos=ctx.user_data["datos"]
        if not datos["recorrido"].get("novedades"):
            nov=novedad_vacia(1); nov["motivo"]=SIN_NOV_MOTIVO; nov["remedio"]=SIN_NOV_REMEDIO
            datos["recorrido"]["novedades"]=[nov]
        xl=generar_excel(datos); nombre=nombre_archivo(datos)
        caption="FOR FO 02 generado"+chr(10)+"Ruta: "+(datos["recorrido"]["nombre_ruta"] or "SIN NOMBRE")+chr(10)+"Novedades: "+str(len(datos["recorrido"]["novedades"]))
        await msg.reply_document(document=xl,filename=nombre,caption=caption)
    except Exception as e:
        logger.error("Error generando Excel: "+str(e))
        try: await msg.reply_text("Error: "+str(e))
        except: pass
    teclado=[["Generar Informe","Nueva Ruta Base"],["Mis Rutas","Ayuda"]]
    try: await msg.reply_text("Que deseas hacer?",reply_markup=ReplyKeyboardMarkup(teclado,resize_keyboard=True))
    except: pass
    return MENU_PRINCIPAL

# ── NUEVA RUTA BASE ───────────────────────────────────────────────────────────
async def nueva_ruta(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id not in USUARIOS_AUTENTICADOS:
        return await start(update, ctx)
    await update.message.reply_text("Nueva Ruta Base"+chr(10)+chr(10)+"Nombre de la ruta:"+chr(10)+"Ejemplo: GOSSEAL-MACHACHI",reply_markup=ReplyKeyboardRemove())
    return NUEVA_RUTA_NOMBRE

async def recv_nueva_ruta_nombre(update, ctx):
    ctx.user_data["nueva_ruta_nombre"]=update.message.text.upper()
    nombre=ctx.user_data["nueva_ruta_nombre"]
    teclado=InlineKeyboardMarkup([[InlineKeyboardButton("Tengo el link de Mapillary",callback_data="vb_link")],[InlineKeyboardButton("Subir video directo aqui",callback_data="vb_video")],[InlineKeyboardButton("Cancelar",callback_data="tab_menu")]])
    await update.message.reply_text("Nueva Ruta Base: "+nombre+chr(10)+chr(10)+"Como quieres registrar el video base?",reply_markup=teclado)
    return NUEVA_RUTA_VIDEO

async def recv_nueva_ruta_video(update, ctx):
    nombre=ctx.user_data.get("nueva_ruta_nombre","SIN NOMBRE")
    if update.message.text and update.message.text.strip().startswith("http"):
        link=update.message.text.strip()
        RUTAS_GUARDADAS[nombre]={"nombre":nombre,"mapillary_link":link,"tipo":"mapillary","fecha":datetime.now().strftime("%d/%m/%Y %H:%M")}
        await update.message.reply_text("✅ Ruta base guardada!"+chr(10)+"Nombre: "+nombre+chr(10)+"Link: "+link)
    elif update.message.video or update.message.document:
        RUTAS_GUARDADAS[nombre]={"nombre":nombre,"tipo":"video_telegram","fecha":datetime.now().strftime("%d/%m/%Y %H:%M")}
        await update.message.reply_text("✅ Video recibido!"+chr(10)+"Ruta base guardada: "+nombre)
    else:
        await update.message.reply_text("Envia el link de Mapillary o el video."+chr(10)+"Ejemplo: https://www.mapillary.com/app/?pKey=xxx")
        return NUEVA_RUTA_VIDEO
    teclado=[["Generar Informe","Nueva Ruta Base"],["Mis Rutas","Ayuda"]]
    await update.message.reply_text("Que deseas hacer?",reply_markup=ReplyKeyboardMarkup(teclado,resize_keyboard=True))
    return MENU_PRINCIPAL

async def vb_callback(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query=update.callback_query; await query.answer(); data=query.data
    if data=="vb_link":
        ctx.user_data["modo_video_base"]="link"
        await query.edit_message_text("Pega el link de Mapillary de tu video base:"+chr(10)+chr(10)+"Ejemplo:"+chr(10)+"https://www.mapillary.com/app/?pKey=ABC123")
        return NUEVA_RUTA_VIDEO
    elif data=="vb_video":
        ctx.user_data["modo_video_base"]="video"
        await query.edit_message_text("Envia el video de la ruta directamente."+chr(10)+"El bot lo subira a Mapillary automaticamente."+chr(10)+"Formatos: .mp4, .mov, .insv")
        return NUEVA_RUTA_VIDEO
    return MENU_PRINCIPAL

# ── MIS RUTAS / AYUDA ─────────────────────────────────────────────────────────
async def mis_rutas(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id not in USUARIOS_AUTENTICADOS:
        return await start(update, ctx)
    if not RUTAS_GUARDADAS:
        await update.message.reply_text("No tienes rutas base guardadas."+chr(10)+"Usa Nueva Ruta Base para registrar.")
        return MENU_PRINCIPAL
    msg="Rutas base registradas:"+chr(10)+chr(10)
    for i,(nombre,info) in enumerate(RUTAS_GUARDADAS.items(),1):
        msg+=str(i)+". "+nombre+chr(10)+"   Fecha: "+info.get("fecha","")+chr(10)
        if info.get("mapillary_link"): msg+="   Link: "+info["mapillary_link"][:50]+"..."+chr(10)
        msg+=chr(10)
    await update.message.reply_text(msg)
    return MENU_PRINCIPAL

async def ayuda(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "RecorridosIA - Ayuda"+chr(10)+chr(10)+
        "Generar Informe — abre el menu de pestanas"+chr(10)+
        "Nueva Ruta Base — registra video base de una ruta"+chr(10)+
        "Mis Rutas — lista rutas guardadas"+chr(10)+chr(10)+
        "Variables en Render:"+chr(10)+"BOT_TOKEN / GEMINI_API_KEY"+chr(10)+"MAPILLARY_TOKEN / TOTP_SECRET / DOMINIO_EMAIL")
    return MENU_PRINCIPAL

async def cancelar(update, ctx):
    ctx.user_data.clear()
    await update.message.reply_text("Cancelado.",reply_markup=ReplyKeyboardRemove())
    return ConversationHandler.END

# ── SERVIDOR WEB ──────────────────────────────────────────────────────────────
class PingHandler(BaseHTTPRequestHandler):
    def do_GET(self):
        self.send_response(200); self.end_headers(); self.wfile.write(b"RecorridosIA OK")
    def log_message(self,format,*args): pass

def ping_render():
    import urllib.request
    while True:
        time.sleep(720)
        try:
            url=os.getenv("RENDER_EXTERNAL_URL","")
            if url: urllib.request.urlopen(url,timeout=10); logger.info("Ping OK - bot despierto")
        except Exception as e: logger.warning("Ping error: "+str(e))

def start_web():
    port=int(os.getenv("PORT",8080))
    server=HTTPServer(("0.0.0.0",port),PingHandler)
    logger.info("Servidor web en puerto "+str(port))
    threading.Thread(target=ping_render,daemon=True).start()
    server.serve_forever()

# ── BUILD APP ─────────────────────────────────────────────────────────────────
def build_app():
    app=Application.builder().token(BOT_TOKEN).build()
    conv=ConversationHandler(
        entry_points=[CommandHandler("start",start),CommandHandler("inspeccionar",generar_informe),MessageHandler(filters.Regex("Generar Informe"),generar_informe),MessageHandler(filters.Regex("Nueva Ruta Base"),nueva_ruta),MessageHandler(filters.Regex("Mis Rutas"),mis_rutas),MessageHandler(filters.Regex("Ayuda"),ayuda)],
        states={
            ESPERANDO_TOTP:   [MessageHandler(filters.TEXT&~filters.COMMAND,handler_totp)],
            MENU_PRINCIPAL:   [MessageHandler(filters.Regex("Generar Informe"),generar_informe),MessageHandler(filters.Regex("Nueva Ruta Base"),nueva_ruta),MessageHandler(filters.Regex("Mis Rutas"),mis_rutas),MessageHandler(filters.Regex("Ayuda"),ayuda),MessageHandler(filters.TEXT&~filters.COMMAND,menu_principal)],
            TAB_MENU:         [MessageHandler(filters.TEXT&~filters.COMMAND,generar_informe)],
            TAB_CIU_HERR:     [MessageHandler(filters.TEXT&~filters.COMMAND,tab_ciu_herr)],
            TAB_CIU_EQUI:     [MessageHandler(filters.TEXT&~filters.COMMAND,tab_ciu_equi)],
            TAB_CIU_MATE:     [MessageHandler(filters.TEXT&~filters.COMMAND,tab_ciu_mate)],
            TAB_MPRIU:        [MessageHandler(filters.TEXT&~filters.COMMAND,tab_mpriu)],
            TAB_REPORTES:     [MessageHandler(filters.TEXT&~filters.COMMAND,tab_reportes)],
            TAB_NOVEDADES_IA: [MessageHandler(filters.PHOTO,tab_novedades_ia),MessageHandler(filters.TEXT&~filters.COMMAND,tab_novedades_ia)],
            MANGA_NOMBRE:     [MessageHandler(filters.TEXT&~filters.COMMAND,recv_manga_nombre)],
            MANGA_COORDS:     [MessageHandler(filters.TEXT&~filters.COMMAND,recv_manga_coords)],
            MANGA_OBS:        [MessageHandler(filters.TEXT&~filters.COMMAND,recv_manga_obs)],
            HILO_ODF:         [MessageHandler(filters.TEXT&~filters.COMMAND,recv_hilo_odf)],
            HILO_DATOS:       [MessageHandler(filters.TEXT&~filters.COMMAND,recv_hilo_datos)],
            NUEVA_RUTA_NOMBRE:[MessageHandler(filters.TEXT&~filters.COMMAND,recv_nueva_ruta_nombre)],
            NUEVA_RUTA_VIDEO: [MessageHandler(filters.TEXT|filters.VIDEO|filters.Document.ALL&~filters.COMMAND,recv_nueva_ruta_video)],
        },
        fallbacks=[CommandHandler("cancelar",cancelar)],
        allow_reentry=True,
    )
    app.add_handler(conv)
    app.add_handler(CallbackQueryHandler(tab_callback,pattern="^tab_"))
    app.add_handler(CallbackQueryHandler(tab_callback,pattern="^rep_"))
    app.add_handler(CallbackQueryHandler(vb_callback,pattern="^vb_"))
    app.add_handler(CallbackQueryHandler(tab_callback,pattern="^manga_der_"))
    return app

async def run_bot():
    app=build_app()
    await app.initialize(); await app.start(); await app.updater.start_polling()
    logger.info("RecorridosIA bot arrancando...")
    while True:
        import asyncio; await asyncio.sleep(1)

def bot_thread():
    import asyncio
    loop=asyncio.new_event_loop(); asyncio.set_event_loop(loop)
    loop.run_until_complete(run_bot())

if __name__=="__main__":
    t=threading.Thread(target=bot_thread,daemon=True); t.start()
    logger.info("RecorridosIA bot arrancando...")
    start_web()
