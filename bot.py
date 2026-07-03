import os, io, hmac, struct, time, base64, hashlib, json, logging, threading
import httpx
from datetime import datetime
from http.server import HTTPServer, BaseHTTPRequestHandler
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from telegram import Update, ReplyKeyboardMarkup, ReplyKeyboardRemove, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Application, CommandHandler, MessageHandler, ConversationHandler, ContextTypes, filters, CallbackQueryHandler

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

BOT_TOKEN       = os.getenv("BOT_TOKEN")
GEMINI_API_KEY  = os.getenv("GEMINI_API_KEY")
MAPILLARY_TOKEN = os.getenv("MAPILLARY_TOKEN")
TOTP_SECRET     = os.getenv("TOTP_SECRET")
DOMINIO         = os.getenv("DOMINIO_EMAIL", "telconet.ec")
GEMINI_URL      = "https://generativelanguage.googleapis.com/v1beta/models/gemini-1.5-flash:generateContent?key=" + (GEMINI_API_KEY or "")

USUARIOS_AUTENTICADOS = set()
RUTAS_GUARDADAS = {}

# MODO PRUEBA: solo se generan REPORTES_DE_RECORRIDOS y FOTOS_ANEXAS_AL_REPORTE.
# Cuando estas 2 pestanas esten validadas, cambiar a True para reactivar
# MANGAS, INVENTARIO DE HILOS EN NODO, Checklist CIU y Checklists MPRIU.
GENERAR_HOJAS_EXTRA = False

(ESPERANDO_TOTP, MENU_PRINCIPAL, NOMBRE_RUTA, CODIGO_CUADRILLA, NODO_INICIAL, NODO_FINAL,
 LIDER, AYUDANTE, COORDINADOR, PLACA, DISTANCIA,
 CIU_HERRAMIENTAS, CIU_EQUIPOS, CIU_MATERIALES,
 NOVEDADES_AUTO, TAREA_PENDIENTE, FOTO_ANTES, FOTO_DESPUES, OBSERVACIONES,
 MPRIU_CHECK, PREGUNTA_MANGAS, PREGUNTA_HILOS,
 MANGA_NOMBRE, MANGA_COORDS, MANGA_OBS, HILO_ODF, HILO_DATOS,
 NUEVA_RUTA_NOMBRE, NUEVA_RUTA_VIDEO,
 GENERAR_CONFIRMAR, GENERAR_NOMBRE_RUTA, GENERAR_CUADRILLA,
 GENERAR_NODO_INI, GENERAR_NODO_FIN, GENERAR_LIDER,
 GENERAR_AYUDANTE, GENERAR_COORDINADOR, GENERAR_PLACA,
 GENERAR_DISTANCIA, GENERAR_FECHA, GENERAR_HORA_INI,
 GENERAR_HORA_FIN, GENERAR_NOVEDADES,
 TAB_MENU, TAB_CIU_HERR, TAB_CIU_EQUI, TAB_CIU_MATE,
 TAB_MPRIU, TAB_REPORTES, TAB_NOVEDADES_IA,
 VIDEO_BASE_NOMBRE, VIDEO_BASE_UPLOAD) = range(52)

NOVEDADES_MPRIU = [
    "HERRAJES EN MAL ESTADO.", "FALTA DE HERRAJES.", "POSTES EN MAL ESTADO.",
    "POSTE(S) CAMBIADO(S).", "POSTES POR INSTALAR.", "POSTE NUEVO INSTALADO \u2013 TN.",
    "POSTE NUEVO INSTALADO \u2013 EMPRESAS EL\u00c9CTRICAS.", "POSTES INCLINADOS.",
    "RETENIDA(S) EN MAL ESTADO.", "RETENIDA(S) CORTADA(S).", "VANOS POR RETEMPLAR.",
    "MANGAS SUELTAS.", "MANGAS ABIERTAS/DA\u00d1ADAS.", "RESERVAS SUELTAS.",
    "CRUCES DE V\u00cdAS BAJOS.", "VEGETACI\u00d3N SOBRE FIBRA/MANGA.", "LOCALIZACI\u00d3N DE MANGA.",
    "DOCUMENTACI\u00d3N UNIFILAR DE HILOS.", "L\u00cdNEA EL\u00c9CTRICA EN MAL ESTADO.",
    "REGENERACI\u00d3N URBANA.", "AMPLIACI\u00d3N DE V\u00cdA.", "CABLE LASTIMADO.",
    "FIBRA INSTALADA INCORRECTAMENTE SOBRE MORDAZA.", "POZO SIN TAPA O EN MAL ESTADO.",
    "REPINTADO DE POZO.", "REPINTADO DE POSTE.", "ELEMENTOS SIN ETIQUETAS ACR\u00cdLICAS.",
    "RIESGO DE DERRUMBE O DESLAVE.", "RIESGO DE INUNDACIONES.", "RIESGO DE INCENDIO.",
    "NO SE REGISTRAN NOVEDADES DURANTE LA INSPECCI\u00d3N.",
]

SIN_NOV_MOTIVO  = "NO SE REGISTRAN NOVEDADES DURANTE LA INSPECCI\u00d3N."
SIN_NOV_REMEDIO = "NO SE ENCUENTRAN NOVEDADES QUE SIGNIFIQUEN RIESGOS EN EL CABLE DE LA RED INTERURBANO."

REMEDIOS = {
    "HERRAJES EN MAL ESTADO.": "REALIZAR EL REEMPLAZO INMEDIATO DEL HERRAJE AFECTADO, GARANTIZANDO LA CORRECTA SUJECI\u00d3N DEL CABLE Y LA ESTABILIDAD MEC\u00c1NICA DEL TENDIDO.",
    "FALTA DE HERRAJES.": "INSTALAR LOS HERRAJES CONFORME A LA NORMATIVA T\u00c9CNICA, ASEGURANDO LA CORRECTA FIJACI\u00d3N DEL CABLE AL POSTE.",
    "POSTES EN MAL ESTADO.": "DOCUMENTAR MEDIANTE REGISTRO FOTOGR\u00c1FICO Y COORDENADAS, Y REPORTAR PARA GESTIONAR EL REEMPLAZO DEL POSTE CON LA ENTIDAD RESPONSABLE.",
    "POSTE(S) CAMBIADO(S).": "INSTALAR LOS HERRAJES NECESARIOS Y ASEGURAR CORRECTAMENTE EL CABLE AL NUEVO POSTE. DOCUMENTAR EL CAMBIO PARA ACTUALIZACI\u00d3N DE INVENTARIO.",
    "POSTES POR INSTALAR.": "DOCUMENTAR LA UBICACI\u00d3N EXACTA Y REPORTAR PARA LA COORDINACI\u00d3N E INSTALACI\u00d3N DEL NUEVO POSTE REQUERIDO.",
    "POSTE NUEVO INSTALADO \u2013 TN.": "DOCUMENTAR, ETIQUETAR CON C\u00d3DIGO DE IDENTIFICACI\u00d3N Y APLICAR PINTURA DE SE\u00d1ALIZACI\u00d3N CONFORME A EST\u00c1NDARES OPERATIVOS.",
    "POSTE NUEVO INSTALADO \u2013 EMPRESAS EL\u00c9CTRICAS.": "DOCUMENTAR, COLOCAR ETIQUETA ACR\u00cdLICA Y ASEGURAR EL CABLE DE FIBRA \u00d3PTICA CONFORME A LA NORMATIVA T\u00c9CNICA VIGENTE, RESPETANDO LAS DISTANCIAS DE SEGURIDAD EL\u00c9CTRICA.",
    "POSTES INCLINADOS.": "DOCUMENTAR MEDIANTE REGISTRO FOTOGR\u00c1FICO Y COORDENADAS, Y REPORTAR PARA GESTIONAR LA CORRECCI\u00d3N CON EL CONTRATISTA.",
    "RETENIDA(S) EN MAL ESTADO.": "DOCUMENTAR MEDIANTE REGISTRO FOTOGR\u00c1FICO Y COORDENADAS, Y REPORTAR PARA GESTIONAR LA CORRECCI\u00d3N CON EL CONTRATISTA.",
    "RETENIDA(S) CORTADA(S).": "DOCUMENTAR MEDIANTE REGISTRO FOTOGR\u00c1FICO Y COORDENADAS, Y REPORTAR PARA GESTIONAR LA CORRECCI\u00d3N CON EL CONTRATISTA.",
    "VANOS POR RETEMPLAR.": "REALIZAR EL RETEMPLADO DEL CABLE PARA RESTABLECER LA TENSI\u00d3N ADECUADA Y EVITAR RIESGOS DE DA\u00d1O O CA\u00cdDA.",
    "MANGAS SUELTAS.": "ASEGURAR LA MANGA AL POSTE EN CONFIGURACI\u00d3N TIPO 'FIGURA 8', CONFORME AL EST\u00c1NDAR.",
    "MANGAS ABIERTAS/DA\u00d1ADAS.": "REEMPLAZAR TAPAS Y SELLOS, GARANTIZANDO EL CIERRE HERM\u00c9TICO Y LA PROTECCI\u00d3N DEL EMPALME CONTRA AGENTES EXTERNOS.",
    "RESERVAS SUELTAS.": "REORGANIZAR Y ASEGURAR LA RESERVA EN 'FIGURA 8' CONFORME A LO ESTABLECIDO.",
    "CRUCES DE V\u00cdAS BAJOS.": "AJUSTAR LA ALTURA DEL CABLE ELEV\u00c1NDOLO A LA DISTANCIA REGLAMENTARIA O REPORTAR PARA LA IMPLEMENTACI\u00d3N DE UNA SOLUCI\u00d3N ESTRUCTURAL.",
    "VEGETACI\u00d3N SOBRE FIBRA/MANGA.": "REALIZAR LA PODA O RETIRO DE VEGETACI\u00d3N QUE COMPROMETA LA INTEGRIDAD O SEGURIDAD DEL CABLE. EN CASO DE REQUERIR PERMISOS, DOCUMENTAR LA NOVEDAD.",
    "LOCALIZACI\u00d3N DE MANGA.": "DOCUMENTAR LA UBICACI\u00d3N MEDIANTE COORDENADAS GPS Y REGISTRO FOTOGR\u00c1FICO PARA ACTUALIZACI\u00d3N DE INVENTARIO.",
    "DOCUMENTACI\u00d3N UNIFILAR DE HILOS.": "DOCUMENTAR O SOLICITAR LA PROGRAMACI\u00d3N DE TRABAJO PARA OBTENER LA INFORMACI\u00d3N; UTILIZAR UN SEGUIDOR DE SE\u00d1AL.",
    "L\u00cdNEA EL\u00c9CTRICA EN MAL ESTADO.": "DOCUMENTAR EL RIESGO Y SOLICITAR AL COORDINADOR EL REPORTE AL \u00c1REA DE REGULATORIO.",
    "REGENERACI\u00d3N URBANA.": "ESTABLECER CONTACTO CON EL CONSORCIO, DOCUMENTAR LA AFECTACI\u00d3N Y COORDINAR, JUNTO CON EL COORDINADOR DE FO, LAS MEDIDAS DE MITIGACI\u00d3N, ASIGNANDO TAREAS A LOS DEPARTAMENTOS COMPETENTES COMO REGULATORIO, OBRA CIVIL Y \u00c1REAS INVOLUCRADAS.",
    "AMPLIACI\u00d3N DE V\u00cdA.": "DOCUMENTAR, REGISTRAR EL CONTACTO DEL RESPONSABLE DE LA OBRA Y COORDINAR MEDIDAS DE MITIGACI\u00d3N DE LA INFRAESTRUCTURA CON EL COORDINADOR DE FO.",
    "CABLE LASTIMADO.": "DOCUMENTAR E INFORMAR PARA PROGRAMAR EL CAMBIO DEL TRAMO DE CABLE.",
    "FIBRA INSTALADA INCORRECTAMENTE SOBRE MORDAZA.": "CORREGIR LA INSTALACI\u00d3N SEPARANDO ADECUADAMENTE EL CABLE DE FIBRA DEL MENSAJERO CONFORME A LA NORMATIVA T\u00c9CNICA.",
    "POZO SIN TAPA O EN MAL ESTADO.": "SOLICITAR LA EJECUCI\u00d3N DE TRABAJOS DE OBRA CIVIL PARA SU INSTALACI\u00d3N O CORRECCI\u00d3N.",
    "REPINTADO DE POZO.": "REALIZAR EL PINTADO DEL POZO TELCONET CON EL C\u00d3DIGO ASIGNADO POR GIS.",
    "REPINTADO DE POSTE.": "REALIZAR EL PINTADO DEL POSTE TELCONET CON EL C\u00d3DIGO ASIGNADO POR GIS.",
    "ELEMENTOS SIN ETIQUETAS ACR\u00cdLICAS.": "VERIFICAR, COLOCAR ETIQUETA ACR\u00cdLICA Y ETIQUETAR CON EL C\u00d3DIGO DE RUTA.",
    "RIESGO DE DERRUMBE O DESLAVE.": "DOCUMENTAR EL RIESGO Y SOLICITAR AL COORDINADOR LA REUBICACI\u00d3N DEL RECORRIDO DEL CABLE.",
    "RIESGO DE INUNDACIONES.": "DOCUMENTAR EL RIESGO Y SOLICITAR AL COORDINADOR LA REUBICACI\u00d3N DEL RECORRIDO DEL CABLE.",
    "RIESGO DE INCENDIO.": "DOCUMENTAR EL RIESGO Y SOLICITAR AL COORDINADOR LA REUBICACI\u00d3N DEL RECORRIDO DEL CABLE.",
}

HERR = ["Cintur\u00f3n y Linea de Vida","Casco","Escalera de 24 pies","Escalera de 28 pies","Escalera de 32 pies","Conos reflectivos","Caja para herramientas","Juego de destornilladores","Martillo mediano","Estiletes","Cortafr\u00edo","Alicate","Llave francesa","Juego de rachet","Pares de guantes aislantes","Tecle","Machete","Cizalla","Pata de cabra","Flejadora (M\u00e1quina Eriband)","Extensi\u00f3n con foco","Motosierra","Tijeras met\u00e1licas","Arco de sierra","Binoculares","Parasol","Remolque / Carrete para F.O."]
EQUI = ["Fusionadora","Cortadora de fibra","Bobina de lanzamiento","OTDR con cargador","Llave Acsys","GPS","Inversor","Etiquetadora"]
MATE = ["Fibra 48h (500mt)","Mangas de 48h y/o 144h (2 m\u00ednimo)","Rollo de cinta Eriband 3/4\"","Hebillas para cinta Eriband 3/4\"","Hojas de sierra","Patchcord de fibra","Adaptadores (Simplex-Duplex)","Paquetes de amarras","Mesas pl\u00e1sticas","Sillas pl\u00e1sticas","Cuchillos","Poleas","Sogas de nylon medianas","Sogas de nylon gruesas","Repelente contra insectos","Repelente contra abejas y avispas"]

LOGO_B64 = "iVBORw0KGgoAAAANSUhEUgAAAqcAAACoCAYAAAGqK3duAAAAAXNSR0IArs4c6QAAAARnQU1BAACxjwv8YQUAAAAJcEhZcwAAIdUAACHVAQSctJ0AACTYSURBVHhe7d0NkBxnfefxljDGYCAYCBxg7CPxG7IlWZqZlS3JXmu1uzMt2ZYsTffaWLLedmZn15J2pxdXpY4iDrnUcUnq6iAXwhXhOJJwBK54uQS4I7niOIKPCpfwfgFycEBMyhAOcPDZBozMXj+zz7P7TM9/ZrpnemZndr+fql9pp/vpp3"

SOLUCIONES = REMEDIOS

AZUL   = "002060"
BLANCO = "FFFFFF"
GRIS   = "F2F2F2"
GRIS2  = "EAEAEA"
GRIS3  = "D9D9D9"
VERDE  = "C6EFCE"
ROJO   = "FFC7CE"

_THIN  = Side(style='thin', color='000000')
_BORDE = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_BORDE_GRUESO = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))

def _borde_celda(ws, fila, col, left=_THIN, right=_THIN, top=_THIN, bottom=_THIN):
    ws.cell(fila, col).border = Border(left=left, right=right, top=top, bottom=bottom)

def _hdr(ws, fila, col_ini, col_fin, texto, fg=AZUL, color=BLANCO, bold=True, size=11, borde=None):
    from openpyxl.styles import Alignment
    if col_fin > col_ini:
        ws.merge_cells(start_row=fila, start_column=col_ini, end_row=fila, end_column=col_fin)
    b = borde if borde else _BORDE
    for cc in range(col_ini, col_fin+1):
        cell = ws.cell(fila, cc)
        cell.border = b
    c = ws.cell(fila, col_ini, texto)
    c.font = Font(bold=bold, name="Calibri", size=size, color=color)
    c.fill = PatternFill("solid", fgColor=fg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return c

def _lbl(ws, fila, col, texto, bg=GRIS2, bold=True):
    from openpyxl.styles import Alignment
    c = ws.cell(fila, col, texto)
    c.font = Font(bold=bold, name="Calibri", size=11, color="000000")
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.border = _BORDE
    return c

def _val(ws, fila, col_ini, col_fin, texto, bold=False, borde=None):
    from openpyxl.styles import Alignment
    if col_fin > col_ini:
        ws.merge_cells(start_row=fila, start_column=col_ini, end_row=fila, end_column=col_fin)
    b = borde if borde else _BORDE
    for cc in range(col_ini, col_fin+1):
        cell = ws.cell(fila, cc)
        cell.border = b
    c = ws.cell(fila, col_ini, str(texto))
    c.font = Font(bold=bold, name="Calibri", size=11, color="000000")
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    return c

def _insertar_logo_centrado(ws, fila, col_letra):
    from openpyxl.drawing.image import Image as XLImage
    try:
        data = base64.b64decode(LOGO_B64)
        img = XLImage(io.BytesIO(data))
        img.width = 120
        img.height = 45
        ws.add_image(img, f"{col_letra}{fila}")
    except Exception as e:
        logger.error("Error logo: " + str(e))

async def analizar_con_gemini_ia(img_bytes):
    if not GEMINI_API_KEY:
        return None
    img_b64 = base64.b64encode(img_bytes).decode()
    prompt = "Analiza esta imagen de inspeccion de fibra optica. Si hay problema responde JSON: {\"tiene_novedad\": true, \"motivo\": \"NOMBRE EN MAYUSCULAS\", \"coordenadas\": \"\"}. Si todo bien: {\"tiene_novedad\": false, \"motivo\": \"\", \"coordenadas\": \"\"}. Solo JSON."
    payload = {"contents": [{"parts": [{"text": prompt}, {"inline_data": {"mime_type": "image/jpeg", "data": img_b64}}]}]}
    try:
        async with httpx.AsyncClient(timeout=30) as client:
            resp = await client.post(GEMINI_URL, json=payload)
            texto = resp.json()["candidates"][0]["content"]["parts"][0]["text"]
            texto = texto.replace("```json","").replace("```","").strip()
            r = json.loads(texto)
            if not r.get("tiene_novedad"):
                return None
            motivo = r.get("motivo","").upper()
            return {"motivo": motivo, "remedio": REMEDIOS.get(motivo, "DOCUMENTAR Y REPORTAR AL COORDINADOR."), "coordenadas": r.get("coordenadas",""), "tarea_pendiente": "", "foto_antes": img_bytes, "foto_despues": None}
    except Exception as e:
        logger.error("Gemini error: " + str(e))
        return None

def _logo_image():
    from openpyxl.drawing.image import Image as XLImage
    try:
        data = base64.b64decode(LOGO_B64)
        img = XLImage(io.BytesIO(data))
        img.width = 110
        img.height = 42
        return img
    except:
        return None

def generar_excel_bytes(datos):
    wb = Workbook()
    r = datos["recorrido"]
    ciu = datos["ciu"]
    nch = datos["mpriu"]["novedades_check"]

    ws1 = wb.active
    ws1.title = "REPORTES_DE_RECORRIDOS"
    ws1.views.sheetView[0].showGridLines = True
    for col, w in [("A", 35), ("B", 25), ("C", 20), ("D", 20)]:
        ws1.column_dimensions[col].width = w

    ws1.row_dimensions[2].height = 60
    ws1.cell(2, 1).border = _BORDE_GRUESO
    _insertar_logo_centrado(ws1, 2, "A")
    _hdr(ws1, 2, 2, 3, "REPORTE DE RECORRIDOS DE MANTENIMIENTO PREVENTIVO PARA RUTAS INTERURBANAS", fg="FFFFFF", color="000000", borde=_BORDE_GRUESO)
    _val(ws1, 2, 4, 4, "Código: FOR FO 02\nVersión: 3 (28/05/2021)", bold=True, borde=_BORDE_GRUESO)

    ws1.row_dimensions[4].height = 20
    _hdr(ws1, 4, 1, 4, "DATOS INFORMATIVOS")
    for f, lbl, val in [(5, "Nombre de Ruta", r.get("nombre_ruta","")), (6, "Código de Cuadrilla", r.get("codigo_cuadrilla","")), (7, "Tramo (Nodo Inicial - Nodo Final)", f"{r.get('nodo_inicial','')} - {r.get('nodo_final','')}")]:
        ws1.row_dimensions[f].height = 25
        _lbl(ws1, f, 1, lbl)
        _val(ws1, f, 2, 4, val)

    ws1.row_dimensions[8].height = 20
    _hdr(ws1, 8, 1, 4, "REGISTRO DE NOVEDADES EN LA RUTA")

    fila = 10
    novedades = r.get("novedades", [])[:20]
    for nov in novedades:
        num = str(nov.get("numero", ""))
        ws1.row_dimensions[fila].height = 38
        ws1.merge_cells(start_row=fila, start_column=1, end_row=fila+1, end_column=1)
        for fr in (fila, fila+1):
            ws1.cell(fr, 1).border = _BORDE
        _lbl(ws1, fila, 1, "FECHA Y HORA NOVEDAD # "+num, bg=GRIS)
        ws1.cell(fila, 1).alignment = ws1.cell(fila, 1).alignment.copy(horizontal="center")
        _lbl(ws1, fila, 2, "FECHA"); _lbl(ws1, fila, 3, "HORA INICIO"); _lbl(ws1, fila, 4, "HORA FIN")
        fila += 1; ws1.row_dimensions[fila].height = 38
        _val(ws1, fila, 2, 2, nov.get("fecha", "")); _val(ws1, fila, 3, 3, nov.get("hora_inicio", "")); _val(ws1, fila, 4, 4, nov.get("hora_fin", ""))
        fila += 1
        for label, key in [("MOTIVO APARENTE DE LA NOVEDAD", "motivo"), ("REMEDIO DEFINITIVO A LA NOVEDAD", "remedio"), ("TAREA PENDIENTE (por regulatorio/obra civil, contratista)", "tarea_pendiente"), ("COORDENADAS SITIO DE LA NOVEDAD (Grados decimales)", "coordenadas")]:
            ws1.row_dimensions[fila].height = 42
            _lbl(ws1, fila, 1, label, bg=GRIS3); _val(ws1, fila, 2, 4, nov.get(key, ""))
            fila += 1
        fila += 1

    ws2 = wb.create_sheet("FOTOS_ANEXAS_AL_REPORTE")
    ws2.views.sheetView[0].showGridLines = True
    for col, w in [("A", 4), ("B", 35), ("C", 40), ("D", 40), ("E", 35)]:
        ws2.column_dimensions[col].width = w
    ws2.row_dimensions[2].height = 60
    ws2.cell(2, 2).border = _BORDE_GRUESO
    _insertar_logo_centrado(ws2, 2, "A")
    _hdr(ws2, 2, 3, 4, "REPORTE DE RECORRIDOS DE MANTENIMIENTO PREVENTIVO PARA RUTAS INTERURBANAS", fg="FFFFFF", color="000000", borde=_BORDE_GRUESO)
    _val(ws2, 2, 5, 5, "Código: FOR FO 02\nVersión: 3 (28/05/2021)", bold=True, borde=_BORDE_GRUESO)
    ws2.row_dimensions[4].height = 18
    _hdr(ws2, 4, 2, 5, "FOTOS DE LAS ACCIONES CORRECTIVAS", borde=_BORDE_GRUESO)
    ws2.row_dimensions[5].height = 4; ws2.row_dimensions[6].height = 4
    for fr in (5, 6):
        ws2.cell(fr, 2).border = Border(left=_THIN)
        ws2.cell(fr, 5).border = Border(right=_THIN)

    from openpyxl.styles import Alignment as _Al
    ws2.merge_cells(start_row=7, start_column=2, end_row=8, end_column=2)
    _lbl(ws2, 7, 2, f"NODO DE INICIO:\n{r.get('nodo_inicial','')}", bg="D9E1F2")
    ws2.cell(7, 2).alignment = _Al(horizontal="center", vertical="center", wrap_text=True)
    _hdr(ws2, 7, 3, 3, "FOTO ANTES DE LA ACCIÓN CORRECTIVA", fg="1F4E79")
    _hdr(ws2, 7, 4, 4, "FOTO DESPUÉS DE LA ACCIÓN CORRECTIVA", fg="1F4E79")
    ws2.merge_cells(start_row=7, start_column=5, end_row=8, end_column=5)
    _lbl(ws2, 7, 5, f"NODO FINAL:\n{r.get('nodo_final','')}", bg="D9E1F2")
    ws2.cell(7, 5).alignment = _Al(horizontal="center", vertical="center", wrap_text=True)

    ws2.row_dimensions[8].height = 250
    _val(ws2, 8, 3, 3, "[Foto Antes]"); _val(ws2, 8, 4, 4, "[Foto Después]")
    ws2.cell(8, 3).alignment = _Al(horizontal="center", vertical="center")
    ws2.cell(8, 4).alignment = _Al(horizontal="center", vertical="center")

    from openpyxl.drawing.image import Image as XLImage
    idx_img = 1
    for nov in novedades:
        if nov.get("foto_antes"):
            try:
                img_a = XLImage(io.BytesIO(nov["foto_antes"]))
                img_a.width, img_a.height = 280, 240
                ws2.add_image(img_a, f"C{8}")
            except Exception as ex:
                logger.error("Error foto_antes Excel: "+str(ex))
        if nov.get("foto_despues"):
            try:
                img_d = XLImage(io.BytesIO(nov["foto_despues"]))
                img_d.width, img_d.height = 280, 240
                ws2.add_image(img_d, f"D{8}")
            except Exception as ex:
                logger.error("Error foto_despues Excel: "+str(ex))
        break

    if not GENERAR_HOJAS_EXTRA:
        buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf.read()

    ws3 = wb.create_sheet("MANGAS")
    ws3.views.sheetView[0].showGridLines = True
    for col, w in [("A", 4), ("B", 18), ("C", 40), ("D", 18), ("E", 40)]:
        ws3.column_dimensions[col].width = w
    ws3.row_dimensions[2].height = 46
    _hdr(ws3, 2, 2, 5, "MANGAS DE LA RUTA")
    ws3.row_dimensions[4].height = 18
    _hdr(ws3, 4, 2, 5, "FOTOS DE LAS MANGAS DESDE EL NODO A AL B", fg="00133A")
    mangas = datos.get("mangas", [])[:40]
    f3 = 6
    n_pares = max(20, (len(mangas)+1)//2)
    for i in range(0, n_pares*2, 2):
        m1 = mangas[i] if i < len(mangas) else {}
        m2 = mangas[i+1] if i+1 < len(mangas) else {}
        _hdr(ws3, f3, 2, 2, "NOMBRE:", fg="1F4E79", size=10)
        _val(ws3, f3, 3, 3, m1.get("nombre", ""))
        _hdr(ws3, f3, 4, 4, "NOMBRE:", fg="1F4E79", size=10)
        _val(ws3, f3, 5, 5, m2.get("nombre", "")); f3 += 1
        ws3.row_dimensions[f3].height = 315
        _val(ws3, f3, 3, 3, "")
        _val(ws3, f3, 5, 5, ""); f3 += 1
        for label, k in [("DERIVACI\u00d3N:", "derivacion"), ("COORDENADAS:", "coordenadas"), ("OBSERVACI\u00d3N:", "observacion")]:
            _hdr(ws3, f3, 2, 2, label, fg="1F4E79", size=10)
            _val(ws3, f3, 3, 3, m1.get(k, ""))
            _hdr(ws3, f3, 4, 4, label, fg="1F4E79", size=10)
            _val(ws3, f3, 5, 5, m2.get(k, ""))
            f3 += 1
        f3 += 1

    ws4 = wb.create_sheet("INVENTARIO DE HILOS EN NODO")
    ws4.views.sheetView[0].showGridLines = True
    for col, w in [("A", 11), ("B", 11), ("C", 40), ("D", 4), ("E", 11), ("F", 11), ("G", 40)]:
        ws4.column_dimensions[col].width = w
    ws4.row_dimensions[2].height = 46
    _hdr(ws4, 2, 1, 7, "REGISTRO DE INVENTARIO DE HILOS EN NODO")
    _lbl(ws4, 4, 1, "POSICIÓN ODF:", bg="GRIS")
    _val(ws4, 4, 2, 3, datos["hilos"].get("posicion_odf", ""))
    _hdr(ws4, 6, 1, 1, "PAR"); _hdr(ws4, 6, 2, 2, "HILO"); _hdr(ws4, 6, 3, 3, "DESCRIPCIÓN / SERVICIO")
    _hdr(ws4, 6, 5, 5, "PAR"); _hdr(ws4, 6, 6, 6, "HILO"); _hdr(ws4, 6, 7, 7, "DESCRIPCIÓN / SERVICIO")
    filas_h = datos["hilos"].get("filas", [])
    b1, b2 = filas_h[:24], filas_h[24:48]
    for idx in range(24):
        f4 = 7 + idx
        ws4.row_dimensions[f4].height = 20
        p = str(idx // 2 + 1)
        h1 = b1[idx] if idx < len(b1) else {}
        _val(ws4, f4, 1, 1, p); _val(ws4, f4, 2, 2, str(idx+1)); _val(ws4, f4, 3, 3, h1.get("descripcion", ""))
        h2 = b2[idx] if idx < len(b2) else {}
        _val(ws4, f4, 5, 5, str(idx // 2 + 13)); _val(ws4, f4, 6, 6, str(idx+25)); _val(ws4, f4, 7, 7, h2.get("descripcion", ""))

    ws5 = wb.create_sheet("Checklist CIU")
    for col, w in [("A", 9), ("B", 26), ("C", 11), ("D", 21), ("E", 11), ("F", 14), ("G", 11), ("H", 14)]:
        ws5.column_dimensions[col].width = w
    ws5.row_dimensions[2].height = 46
    _hdr(ws5, 2, 2, 7, "CHECKLIST CUADRILLA INTERURBANA")
    _val(ws5, 2, 8, 8, "Código: FOR FO 05\nVersión: 3 (26/06/2025)", bold=True)
    _lbl(ws5, 4, 2, "Fecha del Recorrido", bg=None); _val(ws5, 4, 3, 4, r.get("fecha", ""))
    _lbl(ws5, 4, 5, "Hora Inicio", bg=None); _val(ws5, 4, 6, 6, r.get("hora_inicio", ""))
    _lbl(ws5, 4, 7, "Hora Fin", bg=None); _val(ws5, 4, 8, 8, r.get("hora_fin", ""))
    f5 = 5
    for label, valor in [("Nombre de Ruta", r.get("nombre_ruta", "")), ("Nodo Inicio", r.get("nodo_inicial", "")), ("Nodo Final", r.get("nodo_final", "")), ("Distancia de la Ruta", ciu.get("distancia_ruta", "")), ("Lider de Cuadrilla", r.get("lider", "")), ("Ayudante de Cuadrilla", r.get("ayudante", "")), ("Coordinador de FO", r.get("coordinador", "")), ("Placa de Vehículo", ciu.get("vehiculo_placa", ""))]:
        _lbl(ws5, f5, 2, label, bg=None); _val(ws5, f5, 3, 8, valor); f5 += 1
    f5 += 1
    _hdr(ws5, f5, 2, 2, "HERRAMIENTAS / EPP"); _hdr(ws5, f5, 3, 3, "CANT"); _hdr(ws5, f5, 4, 4, "OBS")
    _hdr(ws5, f5, 5, 5, "EQUIPOS ELECTRÓNICOS"); _hdr(ws5, f5, 6, 6, "CANT"); _hdr(ws5, f5, 7, 7, "OBS")
    _hdr(ws5, f5, 8, 8, "ESTADO GENERADOR", fg="548235")
    f5 += 1
    ch, ce, cm = ciu.get("herramientas", {}), ciu.get("equipos", {}), ciu.get("materiales", {})
    for i in range(max(len(HERR), len(EQUI), len(MATE))):
        ws5.row_dimensions[f5].height = 20
        if i < len(HERR):
            n = HERR[i]; info = ch.get(n, {"cantidad": 0, "obs": "NINGUNA"})
            _val(ws5, f5, 2, 2, n); _val(ws5, f5, 3, 3, info["cantidad"]); _val(ws5, f5, 4, 4, info["obs"])
        if i < len(EQUI):
            n = EQUI[i]; info = ce.get(n, {"cantidad": 0, "obs": "NINGUNA"})
            _val(ws5, f5, 5, 5, n); _val(ws5, f5, 6, 6, info["cantidad"]); _val(ws5, f5, 7, 7, info["obs"])
        if i == 0:
            _val(ws5, f5, 8, 8, "BUENO")
        f5 += 1

    ws6 = wb.create_sheet("Checklists MPRIU")
    for col, w in [("A", 4), ("B", 42), ("C", 10), ("D", 22), ("E", 22), ("F", 22), ("G", 22), ("H", 11)]:
        ws6.column_dimensions[col].width = w
    ws6.row_dimensions[2].height = 46
    _hdr(ws6, 2, 2, 7, "CHECKLIST DE INSPECCIÓN DE MANTENIMIENTO PREVENTIVO DE LA RED INTERURBANA (MPRIU)")
    _val(ws6, 2, 8, 8, "Código: FOR FO 06\nVersión: 3 (26/06/2025)", bold=True)
    _hdr(ws6, 4, 2, 2, "ACTIVIDADES / NOVEDADES COMPLEMENTARIAS A VERIFICAR EN LA INSPECCIÓN"); _hdr(ws6, 4, 3, 3, "ESTADO"); _hdr(ws6, 4, 4, 7, "SOLUCIONES SUGERIDAS"); _hdr(ws6, 4, 8, 8, "CANTIDAD")
    f6 = 5
    for novedad in NOVEDADES_MPRIU:
        ws6.row_dimensions[f6].height = 43
        info = nch.get(novedad, {}); tiene = info.get("check", False); cant = info.get("cantidad", 0); chk = "SI" if tiene else "NO"
        sol = SOLUCIONES.get(novedad, "DOCUMENTAR Y REPORTAR AL COORDINADOR.")
        _val(ws6, f6, 2, 2, novedad)
        _hdr(ws6, f6, 3, 3, chk, fg=(VERDE if tiene else ROJO))
        _val(ws6, f6, 4, 7, sol)
        ws6.cell(f6, 8, cant if tiene else 0).alignment = ws6.cell(f6, 2).alignment.copy(horizontal="center")
        ws6.cell(f6, 8).border = _BORDE
        f6 += 1
    ws6.row_dimensions[f6].height = 60
    _lbl(ws6, f6, 2, "Observaciones:", bg=None); _val(ws6, f6, 3, 8, r.get("observaciones", ""))
    buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf.read()

def datos_vacios():
    return {
        "recorrido": {"fecha": "", "hora_inicio": "", "hora_fin": "", "nombre_ruta": "", "codigo_cuadrilla": "", "nodo_inicial": "", "nodo_final": "", "lider": "", "ayudante": "", "coordinador": "", "fotos_total": 0, "observaciones": "", "novedades": []},
        "ciu": {"vehiculo_placa": "", "distancia_ruta": "", "herramientas": {}, "equipos": {}, "materiales": {}},
        "mpriu": {"novedades_check": {}, "observaciones": ""},
        "mangas": [], "hilos": {"posicion_odf": "", "filas": []},
    }

def novedad_vacia(numero):
    ahora = datetime.now()
    return {"numero": numero, "fecha": ahora.strftime("%d/%m/%Y"), "hora_inicio": ahora.strftime("%H:%M:%S"), "hora_fin": ahora.strftime("%H:%M:%S"), "motivo": "", "remedio": "", "tarea_pendiente": "", "coordenadas": "", "foto_antes": None, "foto_despues": None}

def nombre_archivo(datos):
    ruta = datos["recorrido"]["nombre_ruta"].split()[0].replace("/", "-") if datos["recorrido"]["nombre_ruta"] else "RUTA"
    return "FOR_FO_02_"+ruta+"_"+datetime.now().strftime("%Y%m%d_%H%M")+".xlsx"

# ── AUTENTICACION ─────────────────────────────────────────────────────────────
async def start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id in USUARIOS_AUTENTICADOS:
        return await menu_principal(update, ctx)
    await update.message.reply_text(
        "RecorridosIA — Sistema Inteligente\n\nPor seguridad ingrese el token OTP asignado por su administrador:"
    )
    return ESPERANDO_TOTP

async def verificar_totp(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    token = update.message.text.strip()
    if not TOTP_SECRET:
        USUARIOS_AUTENTICADOS.add(update.effective_user.id)
        return await menu_principal(update, ctx)
    try:
        import hmac, hashlib, time, struct
        key = base64.b32decode(TOTP_SECRET, casefold=True)
        intervals_no = int(time.time() // 30)
        msg = struct.pack(">Q", intervals_no)
        h = hmac.new(key, msg, hashlib.sha1).digest()
        o = h[19] & 15
        h = (struct.unpack(">I", h[o:o+4])[0] & 0x7fffffff) % 1000000
        valido = f"{h:06d}"
        if token == valido:
            USUARIOS_AUTENTICADOS.add(update.effective_user.id)
            return await menu_principal(update, ctx)
    except:
        pass
    await update.message.reply_text("❌ Token inválido o expirado. Intente nuevamente:")
    return ESPERANDO_TOTP

async def menu_principal(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    teclado = [["Nuevo Recorrido Interurbano", "Nueva Ruta Base"], ["Mis Rutas", "Ayuda"]]
    msg = update.message or update.callback_query.message
    await msg.reply_text("RecorridosIA — Menú principal", reply_markup=ReplyKeyboardMarkup(teclado, resize_keyboard=True))
    return MENU_PRINCIPAL

# ── FLUJO INTERURBANO RECORRIDO ──────────────────────────────────────────────
async def nuevo_recorrido(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id not in USUARIOS_AUTENTICADOS:
        return await start(update, ctx)
    ctx.user_data["datos"] = datos_vacios()
    await update.message.reply_text("Módulo Registro de Recorrido Interurbano\n\nNombre de la Ruta:", reply_markup=ReplyKeyboardRemove())
    return NOMBRE_RUTA

async def recv_nombre_ruta(update, ctx):
    ctx.user_data["datos"]["recorrido"]["nombre_ruta"] = update.message.text.upper()
    await update.message.reply_text("Código de Cuadrilla (Ej: C-01):")
    return CODIGO_CUADRILLA

async def recv_codigo_cuadrilla(update, ctx):
    ctx.user_data["datos"]["recorrido"]["codigo_cuadrilla"] = update.message.text.upper()
    await update.message.reply_text("Nodo Inicial:")
    return NODO_INICIAL

async def recv_nodo_inicial(update, ctx):
    ctx.user_data["datos"]["recorrido"]["nodo_inicial"] = update.message.text.upper()
    await update.message.reply_text("Nodo Final:")
    return NODO_FINAL

async def recv_nodo_final(update, ctx):
    ctx.user_data["datos"]["recorrido"]["nodo_final"] = update.message.text.upper()
    await update.message.reply_text("Nombre del Líder de Cuadrilla:")
    return LIDER

async def recv_lider(update, ctx):
    ctx.user_data["datos"]["recorrido"]["lider"] = update.message.text.upper()
    await update.message.reply_text("Nombre del Ayudante:")
    return AYUDANTE

async def recv_ayudante(update, ctx):
    ctx.user_data["datos"]["recorrido"]["ayudante"] = update.message.text.upper()
    await update.message.reply_text("Nombre del Coordinador de FO:")
    return COORDINADOR

async def recv_coordinador(update, ctx):
    ctx.user_data["datos"]["recorrido"]["coordinador"] = update.message.text.upper()
    await update.message.reply_text("Placa del Vehículo:")
    return PLACA

async def recv_placa(update, ctx):
    ctx.user_data["datos"]["ciu"]["vehiculo_placa"] = update.message.text.upper()
    await update.message.reply_text("Distancia de la Ruta (Ej: 15.4 KM):")
    return DISTANCIA

async def recv_distancia(update, ctx):
    ctx.user_data["datos"]["ciu"]["distancia_ruta"] = update.message.text.upper()
    ctx.user_data["datos"]["recorrido"]["fecha"] = datetime.now().strftime("%d/%m/%Y")
    ctx.user_data["datos"]["recorrido"]["hora_inicio"] = datetime.now().strftime("%H:%M:%S")
    return await tab_menu(update, ctx)

# ── PESTAÑAS DINAMICAS DEL INFORME ──────────────────────────────────────────
async def tab_menu(update, ctx):
    datos = ctx.user_data["datos"]
    r = datos["recorrido"]
    c = datos["ciu"]
    m = datos["mpriu"]
    cant_nov = len(r.get("novedades", []))
    cant_mangas = len(datos.get("mangas", []))
    cant_hilos = len(datos["hilos"].get("filas", []))

    chk1 = "✅" if c.get("herramientas") else "❌"
    chk2 = "✅" if c.get("equipos") else "❌"
    chk3 = "✅" if c.get("materiales") else "❌"
    chk4 = "✅" if m.get("novedades_check") else "❌"
    chk5 = "✅" if cant_nov > 0 else "⬜"

    msg = (f"📋 <b>ESTADO DEL INFORME ACTUAL</b>\n"
           f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
           f"Ruta: {r.get('nombre_ruta','')}\n"
           f"Novedades detectadas: {cant_nov}\n"
           f"Mangas guardadas: {cant_mangas}\n"
           f"Hilos registrados: {cant_hilos}\n\n"
           f"Complete los bloques utilizando las opciones de abajo:")

    botones = [
        [InlineKeyboardButton(f"{chk1} Herramientas CIU", callback_data="tab_1"),
         InlineKeyboardButton(f"{chk2} Equipos CIU", callback_data="tab_2")],
        [InlineKeyboardButton(f"{chk3} Materiales CIU", callback_data="tab_3"),
         InlineKeyboardButton(f"{chk4} Check MPRIU", callback_data="tab_4")],
        [InlineKeyboardButton(f"{chk5} Carga Manual / Fotos", callback_data="tab_5"),
         InlineKeyboardButton("🤖 Asistente IA (Fotos)", callback_data="tab_6")]
    ]
    if GENERAR_HOJAS_EXTRA:
        botones.append([InlineKeyboardButton("➕ Agregar Manga", callback_data="tab_manga_add"),
                        InlineKeyboardButton("🎛️ Registrar Hilos ODF", callback_data="tab_hilos")])
    botones.append([InlineKeyboardButton("💾 GENERAR EXCEL (.XLSX)", callback_data="tab_generar")])

    teclado = InlineKeyboardMarkup(botones)
    msg_obj = update.message or update.callback_query.message
    if update.callback_query:
        await update.callback_query.answer()
        try:
            await update.callback_query.edit_message_text(msg, reply_markup=teclado, parse_mode="HTML")
        except:
            await update.callback_query.message.reply_text(msg, reply_markup=teclado, parse_mode="HTML")
    else:
        await update.message.reply_text(msg, reply_markup=teclado, parse_mode="HTML")
    return TAB_MENU

async def tab_callback(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query; await query.answer(); data = query.data
    volver = InlineKeyboardMarkup([[InlineKeyboardButton("Volver al menú", callback_data="tab_menu")]])
    if data == "tab_generar":
        return await enviar_excel(update, ctx)
    elif data == "tab_menu":
        return await tab_menu(update, ctx)
    elif data == "tab_1":
        msg = ("CHECKLIST CIU - HERRAMIENTAS Y EPP\n"
               "━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
               "Escribe cantidades separadas por coma:\n\n"
               " 1. Cinturón y Línea de Vida\n 2. Casco\n 3. Escalera 24 pies\n"
               " 4. Escalera 28 pies\n 5. Escalera 32 pies\n 6. Conos reflectivos\n"
               " 7. Caja para herramientas\n 8. Juego destornilladores\n 9. Martillo mediano\n"
               "10. Estiletes\n11. Cortafrío\n12. Alicate\n13. Llave francesa\n"
               "14. Juego de rachet\n15. Guantes aislantes (pares)\n16. Tecle\n"
               "17. Machete\n18. Cizalla\n19. Pata de cabra\n20. Flejadora Eriband\n"
               "21. Extensión con foco\n22. Motosierra\n23. Tijeras metálicas\n"
               "24. Arco de sierra\n25. Binoculares\n26. Parasol\n27. Remolque/Carrete FO\n\n"
               "Ejemplo: 2,2,0,2,0,6,0,1,1,2,2,0...")
        await query.edit_message_text(msg, reply_markup=volver)
        return TAB_CIU_HERR
    elif data == "tab_2":
        msg = ("CHECKLIST CIU - EQUIPOS ELECTRÓNICOS\n"
               "━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
               "Escribe cantidades separadas por coma:\n\n"
               " 1. Fusionadora\n 2. Cortadora de fibra\n 3. Bobina de lanzamiento\n"
               " 4. OTDR con cargador\n 5. Llave Acsys\n 6. GPS\n"
               " 7. Inversor\n 8. Etiquetadora\n\n"
               "Ejemplo: 1,1,0,1,1,1,0,1")
        await query.edit_message_text(msg, reply_markup=volver)
        return TAB_CIU_EQUI
    elif data == "tab_3":
        msg = ("CHECKLIST CIU - MATERIALES E INSUMOS\n"
               "━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
               "Escribe cantidades separadas por coma:\n\n"
               " 1. Fibra 48h (500mt)\n 2. Mangas 48h/144h\n 3. Rollo cinta Eriband 3/4\n"
               " 4. Hebillas Eriband 3/4\n 5. Hojas de sierra\n 6. Patchcord de fibra\n"
               " 7. Adaptadores (Simplex-Duplex)\n 8. Paquetes de amarras\n 9. Mesas plásticas\n"
               "10. Sillas plásticas\n11. Cuchillos\n12. Poleas\n13. Sogas nylon medianas\n"
               "14. Sogas nylon gruesas\n15. Repelente insectos\n16. Repelente abejas/avispas\n\n"
               "Ejemplo: 0,2,1,20,2,4,4,1,1,2,2,2,2,1,1,1")
        await query.edit_message_text(msg, reply_markup=volver)
        return TAB_CIU_MATE
    elif data == "tab_4":
        msg = ("CHECKLIST MPRIU — REGISTRO DE INSPECCIÓN\n"
               "━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
               "Para marcar novedades que SÍ aplican en la inspección, marque sus índices separados por comas.\n\n"
               "Las que no escriba se marcarán automáticamente como NO NOVEDAD.\n\n")
        for i, n in enumerate(NOVEDADES_MPRIU[:30]):
            msg += f"{i+1:02d}. {n}\n"
        msg += "\nEjemplo: 3,8,12,15"
        await query.edit_message_text(msg, reply_markup=volver)
        return TAB_MPRIU
    elif data == "tab_5":
        msg = ("📝 CARGA MANUAL FORMATO RAPIDO\n"
               "━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
               "Pegue el bloque de texto con los datos y novedades. Formato admitido:\n\n"
               "FECHA: 25/06/2025\n"
               "HORA_INI: 08:30:00\n"
               "HORA_FIN: 17:00:00\n"
               "NOV: HERRAJES EN MAL ESTADO. | -0.2341, -78.5123\n"
               "NOV: CRUCES DE VÍAS BAJOS. | -0.2355, -78.5110\n"
               "OBS: NINGUNA")
        await query.edit_message_text(msg, reply_markup=volver)
        return TAB_REPORTES
    elif data == "tab_6":
        ctx.user_data["media_inspeccion"] = []
        msg = ("🤖 ASISTENTE DE DETECCIÓN CON IA\n"
               "━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
               "Envíe la(s) foto(s) de la inspección en campo una por una. Al finalizar escriba la palabra LISTO.")
        await query.edit_message_text(msg, reply_markup=volver)
        return TAB_NOVEDADES_IA
    elif data == "tab_manga_add":
        await query.edit_message_text("Nombre o Código de la Manga a registrar:")
        return MANGA_NAME
    elif data == "tab_hilos":
        await query.edit_message_text("Posición en ODF (Ej: ODF-03-BDF-01):")
        return HILO_ODF

async def tab_ciu_herr(update, ctx):
    valores = [v.strip() for v in update.message.text.replace(",", " ").split()]
    herr = {}; resumen = ""
    for i, nombre in enumerate(HERR):
        cant = int(valores[i]) if i < len(valores) and valores[i].isdigit() else 0
        herr[nombre] = {"cantidad": cant, "obs": "BUEN ESTADO" if cant > 0 else "NINGUNA"}
        if cant > 0:
            resumen += f" ✅ {nombre}: {cant}\n"
    ctx.user_data["datos"]["ciu"]["herramientas"] = herr
    await update.message.reply_text(f"✅ HERRAMIENTAS GUARDADAS\n━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n{resumen if resumen else ' Ninguna'}\n━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\nEQUIPOS ELECTRÓNICOS — escribe cantidades:\n\n 1. Fusionadora\n 2. Cortadora de fibra\n 3. Bobina de lanzamiento\n 4. OTDR con cargador\n 5. Llave Acsys\n 6. GPS\n 7. Inversor\n 8. Etiquetadora\n\nEjemplo: 1,2,0,1,1,0,1,1")
    return TAB_CIU_EQUI

async def tab_ciu_equi(update, ctx):
    valores = [v.strip() for v in update.message.text.replace(",", " ").split()]
    equi = {}; resumen = ""
    for i, nombre in enumerate(EQUI):
        cant = int(valores[i]) if i < len(valores) and valores[i].isdigit() else 0
        equi[nombre] = {"cantidad": cant, "obs": "BUEN ESTADO" if cant > 0 else "NINGUNA"}
        if cant > 0:
            resumen += f" ✅ {nombre}: {cant}\n"
    ctx.user_data["datos"]["ciu"]["equipos"] = equi
    await update.message.reply_text(f"✅ EQUIPOS GUARDADOS\n━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n{resumen if resumen else ' Ninguno'}\n━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\nMATERIALES E INSUMOS — escribe cantidades:\n\n 1. Fibra 48h (500mt)\n 2. Mangas 48h/144h\n 3. Rollo cinta Eriband 3/4\n 4. Hebillas Eriband 3/4\n 5. Hojas de sierra\n 6. Patchcord de fibra\n 7. Adaptadores (Simplex-Duplex)\n 8. Paquetes de amarras\n 9. Mesas plásticas\n10. Sillas plásticas\n11. Cuchillos\n12. Poleas\n13. Sogas nylon medianas\n14. Sogas nylon gruesas\n15. Repelente insectos\n16. Repelente abejas/avispas")
    return TAB_CIU_MATE

async def tab_ciu_mate(update, ctx):
    valores = [v.strip() for v in update.message.text.replace(",", " ").split()]
    mate = {}; resumen = ""
    for i, nombre in enumerate(MATE):
        cant = int(valores[i]) if i < len(valores) and valores[i].isdigit() else 0
        mate[nombre] = {"cantidad": cant, "obs": "BUEN ESTADO" if cant > 0 else "NINGUNA"}
        if cant > 0:
            resumen += f" ✅ {nombre}: {cant}\n"
    ctx.user_data["datos"]["ciu"]["materiales"] = mate
    await update.message.reply_text("✅ MATERIALES GUARDADOS.")
    return await tab_menu(update, ctx)

async def tab_mpriu(update, ctx):
    valores = [v.strip() for v in update.message.text.replace(",", " ").split() if v.strip().isdigit()]
    indices = [int(v)-1 for v in valores]
    nch = {}
    for i, nov in enumerate(NOVEDADES_MPRIU[:30]):
        if i in indices:
            nch[nov] = {"check": True, "cantidad": 1}
        else:
            nch[nov] = {"check": False, "cantidad": 0}
    ctx.user_data["datos"]["mpriu"]["novedades_check"] = nch
    await update.message.reply_text("✅ Checklist de Inspección guardado exitosamente.")
    return await tab_menu(update, ctx)

async def tab_reportes(update, ctx):
    texto_m = update.message.text
    datos = ctx.user_data["datos"]
    r = datos["recorrido"]
    lineas = [l.strip() for l in texto_m.split("\n") if l.strip()]
    for l in lineas:
        if ":" not in l: continue
        clave, valor = l.split(":", 1)
        clave = clave.strip().upper(); valor = valor.strip()
        if clave == "FECHA": r["fecha"] = valor
        elif clave == "HORA_INI": r["hora_inicio"] = valor
        elif clave == "HORA_FIN": r["hora_fin"] = valor
        elif clave == "OBS" and valor.upper() != "NINGUNA": r["observaciones"] = valor.upper()
    novedades_nuevas = []
    for l in lineas:
        if l.upper().startswith("NOV:"):
            partes = l[4:].strip().split("|")
            motivo = partes[0].strip().upper()
            coords = partes[1].strip() if len(partes) > 1 else ""
            remedio = REMEDIOS.get(motivo, "DOCUMENTAR Y REPORTAR AL COORDINADOR.")
            nov = novedad_vacia(len(novedades_nuevas)+1)
            nov["motivo"] = motivo; nov["remedio"] = remedio; nov["coordenadas"] = coords
            novedades_nuevas.append(nov)
            datos["mpriu"]["novedades_check"][motivo] = {"check": True, "cantidad": 1}
    if novedades_nuevas:
        r["novedades"] = novedades_nuevas
    elif not r.get("novedades"):
        nov = novedad_vacia(1); nov["motivo"] = SIN_NOV_MOTIVO; nov["remedio"] = SIN_NOV_REMEDIO
        r["novedades"] = [nov]
    await update.message.reply_text(f"✅ Datos guardados!\nRuta: {r.get('nombre_ruta','')}\nNovedades: {len(r.get('novedades',[]))}")
    return await tab_menu(update, ctx)

async def tab_novedades_ia(update, ctx):
    if "media_inspeccion" not in ctx.user_data:
        ctx.user_data["media_inspeccion"] = []
    if update.message.photo:
        foto = await update.message.photo[-1].get_file()
        ctx.user_data["media_inspeccion"].append(bytes(await foto.download_as_bytearray()))
        await update.message.reply_text(f"Foto {len(ctx.user_data['media_inspeccion'])} recibida. Envíe más o escriba LISTO")
        return TAB_NOVEDADES_IA
    if update.message.text and update.message.text.upper() == "LISTO":
        await update.message.reply_text("Analizando con Gemini IA...")
        media = ctx.user_data.get("media_inspeccion", []); novedades = []
        for img in media:
            res = await analizar_con_gemini_ia(img)
            if res:
                res["numero"] = len(novedades) + 1
                novedades.append(res)
                ctx.user_data["datos"]["mpriu"]["novedades_check"][res["motivo"]] = {"check": True, "cantidad": 1}
        if novedades:
            ctx.user_data["datos"]["recorrido"]["novedades"] = novedades
            await update.message.reply_text(f"✅ Inteligencia Artificial completada. Se detectaron {len(novedades)} novedades críticas.")
        else:
            nov = novedad_vacia(1); nov["motivo"] = SIN_NOV_MOTIVO; nov["remedio"] = SIN_NOV_REMEDIO
            ctx.user_data["datos"]["recorrido"]["novedades"] = [nov]
            await update.message.reply_text("✅ IA analizada: No se encontraron novedades de riesgo mecánico o de infraestructura.")
        return await tab_menu(update, ctx)
    return TAB_NOVEDADES_IA

async def enviar_excel(update, ctx):
    datos = ctx.user_data.get("datos")
    msg = update.message or update.callback_query.message
    if not datos:
        await msg.reply_text("No hay un informe activo para compilar.")
        return ConversationHandler.END
    await msg.reply_text("Generando y firmando archivo Excel normado FOR-FO-02...")
    try:
        xls_b = generar_excel_bytes(datos)
        nombre = nombre_archivo(datos)
        xl = io.BytesIO(xls_b); xl.name = nombre
        caption = f"FOR-FO-02 Completado\nFecha: {datos['recorrido']['fecha']}\nCuadrilla: {datos['recorrido']['codigo_cuadrilla']}"
        await msg.reply_document(document=xl, filename=nombre, caption=caption)
    except Exception as e:
        logger.error("Error generando Excel: "+str(e))
        try: await msg.reply_text("Error: "+str(e))
        except: pass
    teclado = [["Generar Informe", "Nueva Ruta Base"], ["Mis Rutas", "Ayuda"]]
    try: await msg.reply_text("¿Qué deseas hacer?", reply_markup=ReplyKeyboardMarkup(teclado, resize_keyboard=True))
    except: pass
    return MENU_PRINCIPAL

# ── NUEVA RUTA BASE ───────────────────────────────────────────────────────────
async def nueva_ruta(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id not in USUARIOS_AUTENTICADOS:
        return await start(update, ctx)
    await update.message.reply_text("Nueva Ruta Base\n\nNombre de la ruta:\nEjemplo: GOSSEAL-MACHACHI", reply_markup=ReplyKeyboardRemove())
    return NUEVA_RUTA_NOMBRE

async def recv_nueva_ruta_nombre(update, ctx):
    ctx.user_data["nueva_ruta_nombre"] = update.message.text.upper()
    nombre = ctx.user_data["nueva_ruta_nombre"]
    teclado = InlineKeyboardMarkup([[InlineKeyboardButton("Tengo el link de Mapillary", callback_data="vb_link")], [InlineKeyboardButton("Subir video directo aquí", callback_data="vb_video")], [InlineKeyboardButton("Cancelar", callback_data="tab_menu")]])
    await update.message.reply_text(f"Nueva Ruta Base: {nombre}\n\n¿Cómo quieres registrar el video base?", reply_markup=teclado)
    return NUEVA_RUTA_VIDEO

async def recv_nueva_ruta_video(update, ctx):
    nombre = ctx.user_data.get("nueva_ruta_nombre", "SIN NOMBRE")
    if update.message.text and update.message.text.strip().startswith("http"):
        link = update.message.text.strip()
        RUTAS_GUARDADAS[nombre] = {"nombre": nombre, "mapillary_link": link, "tipo": "mapillary", "fecha": datetime.now().strftime("%d/%m/%Y %H:%M")}
        await update.message.reply_text(f"✅ Ruta base guardada!\nNombre: {nombre}\nOrigen: Mapillary Link")
        return await menu_principal(update, ctx)
    elif update.message.video or update.message.document:
        doc = update.message.video or update.message.document
        RUTAS_GUARDADAS[nombre] = {"nombre": nombre, "file_id": doc.file_id, "tipo": "telegram", "fecha": datetime.now().strftime("%d/%m/%Y %H:%M")}
        await update.message.reply_text(f"✅ Video Base indexado en servidores de Telegram para {nombre}.")
        return await menu_principal(update, ctx)
    return NUEVA_RUTA_VIDEO

async def vb_callback(update, ctx):
    query = update.callback_query; await query.answer(); data = query.data
    if data == "vb_link":
        await query.message.reply_text("Pegue la URL de la secuencia de Mapillary:")
        return NUEVA_RUTA_VIDEO
    elif data == "vb_video":
        await query.message.reply_text("Suba el archivo de video de la ruta aquí (Formato MP4/MOV):")
        return NUEVA_RUTA_VIDEO

async def mis_rutas(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not RUTAS_GUARDADAS:
        await update.message.reply_text("No hay rutas base registradas en el clúster todavía.")
        return
    msg = "🗺️ <b>RUTAS BASE MAPEA-REGISTRADAS</b>\n━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
    for k, v in RUTAS_GUARDADAS.items():
        msg += f"• <b>{v['nombre']}</b> ({v['tipo']}) - {v['fecha']}\n"
    await update.message.reply_text(msg, parse_mode="HTML")

async def ayuda(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "💡 <b>Guía del Sistema RecorridosIA</b>\n\n"
        "• <b>Comando /preguntar:</b> Resuelve dudas técnicas en campo (Ej: <code>/preguntar distancia de vanos estándar</code>).\n"
        "• <b>Informe FOR-FO-02:</b> Genera el documento normado para Telconet de forma automática ingresando los datos en las pestañas.\n"
        "• <b>Visión Artificial:</b> La pestaña de IA analiza fotos reales de infraestructura y añade los remedios definitivos automáticamente.",
        parse_mode="HTML"
    )

# ── COMANDO PREGUNTAR CORREGIDO ──────────────────────────────────────────────
async def preguntar(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    # Validamos si se proporcionó una pregunta después del comando
    if not ctx.args:
        await update.message.reply_text("Por favor, escribe una pregunta después del comando.\nEjemplo: <code>/preguntar qué es un OTDR</code>", parse_mode="HTML")
        return

    prompt = " ".join(ctx.args).strip()
    
    # Validamos que la API key esté cargada en las variables de entorno
    if not GEMINI_API_KEY:
        await update.message.reply_text("Error: La API Key de Gemini no está configurada en el servidor.")
        return

    # Construimos el payload HTTP oficial para Gemini 1.5 Flash
    payload = {
        "contents": [
            {
                "parts": [
                    {"text": prompt}
                ]
            }
        ]
    }

    try:
        # Enviamos indicación visual al usuario en Telegram
        await ctx.bot.send_chat_action(chat_id=update.effective_chat.id, action="typing")
        
        # Realizamos la llamada HTTP asíncrona hacia Google AI Studio
        async with httpx.AsyncClient(timeout=30) as client:
            resp = await client.post(GEMINI_URL, json=payload)
            
            # Verificamos si la respuesta fue exitosa
            if resp.status_code == 200:
                data = resp.json()
                respuesta_texto = data["candidates"][0]["content"]["parts"][0]["text"]
            else:
                logger.error(f"Error HTTP de Google: {resp.status_code} - {resp.text}")
                respuesta_texto = "No pude obtener respuesta de Gemini ahorita. La API Key podría estar sin cuota o bloqueada."
                
    except Exception as e:
        logger.error("Error en /preguntar: " + str(e))
        respuesta_texto = "Ocurrió un error inesperado al procesar tu pregunta en el servidor del bot."

    await update.message.reply_text(respuesta_texto)

async def cancelar(update, ctx):
    ctx.user_data.clear()
    await update.message.reply_text("Cancelado.", reply_markup=ReplyKeyboardRemove())
    return ConversationHandler.END

# ── SERVIDOR WEB ──────────────────────────────────────────────────────────────
class PingHandler(BaseHTTPRequestHandler):
    def do_GET(self):
        self.send_response(200); self.end_headers(); self.wfile.write(b"RecorridosIA OK")
    def log_message(self, format, *args): pass

def ping_render():
    import urllib.request
    while True:
        time.sleep(720)
        try:
            url = os.getenv("RENDER_EXTERNAL_URL", "")
            if url:
                urllib.request.urlopen(url, timeout=10); logger.info("Ping Render OK")
        except:
            pass

def start_server():
    p = int(os.getenv("PORT", "8080"))
    s = HTTPServer(("0.0.0.0", p), PingHandler)
    logger.info(f"Web server en puerto {p}")
    s.serve_forever()

# ── MAIN RUN ──────────────────────────────────────────────────────────────────
def build_app():
    app = Application.builder().token(BOT_TOKEN).build()
    conv = ConversationHandler(
        entry_points=[
            CommandHandler("start", start),
            MessageHandler(filters.Regex("^Nuevo Recorrido Interurbano$"), nuevo_recorrido),
            MessageHandler(filters.Regex("^Nueva Ruta Base$"), nueva_ruta),
            MessageHandler(filters.Regex("^Mis Rutas$"), mis_rutas),
            MessageHandler(filters.Regex("^Ayuda$"), ayuda),
        ],
        states={
            ESPERANDO_TOTP:   [MessageHandler(filters.TEXT & ~filters.COMMAND, verificar_totp)],
            MENU_PRINCIPAL:   [
                MessageHandler(filters.Regex("^Nuevo Recorrido Interurbano$"), nuevo_recorrido),
                MessageHandler(filters.Regex("^Nueva Ruta Base$"), nueva_ruta),
                MessageHandler(filters.Regex("^Mis Rutas$"), mis_rutas),
                MessageHandler(filters.Regex("^Ayuda$"), ayuda),
            ],
            NOMBRE_RUTA:      [MessageHandler(filters.TEXT & ~filters.COMMAND, recv_nombre_ruta)],
            CODIGO_CUADRILLA: [MessageHandler(filters.TEXT & ~filters.COMMAND, recv_codigo_cuadrilla)],
            NODO_INICIAL:     [MessageHandler(filters.TEXT & ~filters.COMMAND, recv_nodo_inicial)],
            NODO_FINAL:       [MessageHandler(filters.TEXT & ~filters.COMMAND, recv_nodo_final)],
            LIDER:            [MessageHandler(filters.TEXT & ~filters.COMMAND, recv_lider)],
            AYUDANTE:         [MessageHandler(filters.TEXT & ~filters.COMMAND, recv_ayudante)],
            COORDINADOR:      [MessageHandler(filters.TEXT & ~filters.COMMAND, recv_coordinador)],
            PLACA:            [MessageHandler(filters.TEXT & ~filters.COMMAND, recv_placa)],
            DISTANCIA:        [MessageHandler(filters.TEXT & ~filters.COMMAND, recv_distancia)],
            TAB_MENU:         [CallbackQueryHandler(tab_callback, pattern="^tab_")],
            TAB_CIU_HERR:     [MessageHandler(filters.TEXT & ~filters.COMMAND, tab_ciu_herr)],
            TAB_CIU_EQUI:     [MessageHandler(filters.TEXT & ~filters.COMMAND, tab_ciu_equi)],
            TAB_CIU_MATE:     [MessageHandler(filters.TEXT & ~filters.COMMAND, tab_ciu_mate)],
            TAB_MPRIU:        [MessageHandler(filters.TEXT & ~filters.COMMAND, tab_mpriu)],
            TAB_REPORTES:     [MessageHandler(filters.TEXT & ~filters.COMMAND, tab_reportes)],
            TAB_NOVEDADES_IA: [MessageHandler(filters.PHOTO | filters.TEXT & ~filters.COMMAND, tab_novedades_ia)],
            MANGA_NOMBRE:     [MessageHandler(filters.TEXT & ~filters.COMMAND, cancelar)], # Marcadores por defecto
            HILO_ODF:         [MessageHandler(filters.TEXT & ~filters.COMMAND, cancelar)],
            NUEVA_RUTA_NOMBRE:[MessageHandler(filters.TEXT & ~filters.COMMAND, recv_nueva_ruta_nombre)],
            NUEVA_RUTA_VIDEO: [MessageHandler(filters.TEXT | filters.VIDEO | filters.Document.ALL & ~filters.COMMAND, recv_nueva_ruta_video)],
        },
        fallbacks=[CommandHandler("cancelar", cancelar)],
        allow_reentry=True,
    )
    app.add_handler(conv)
    app.add_handler(CommandHandler("preguntar", preguntar))
    app.add_handler(CallbackQueryHandler(tab_callback, pattern="^tab_"))
    app.add_handler(CallbackQueryHandler(tab_callback, pattern="^rep_"))
    app.add_handler(CallbackQueryHandler(vb_callback, pattern="^vb_"))
    app.add_handler(CallbackQueryHandler(tab_callback, pattern="^manga_der_"))
    return app

async def run_bot():
    app = build_app()
    await app.initialize(); await app.start(); await app.updater.start_polling()
    logger.info("RecorridosIA bot arrancando...")
    while True:
        import asyncio; await asyncio.sleep(1)

def bot_thread():
    import asyncio
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    loop.run_until_complete(run_bot())

if __name__ == "__main__":
    if not BOT_TOKEN:
        raise ValueError("Falta BOT_TOKEN en las variables de entorno.")
    threading.Thread(target=start_server, daemon=True).start()
    threading.Thread(target=ping_render, daemon=True).start()
    bot_thread()
